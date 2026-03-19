#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ============================================================
# GitHub Copilot 利用状況サマリーを Excel に1行追記するスクリプト
#
# 【このスクリプトがやること】
# - user.csv（氏名、GitHubユーザー名、参加区分、Copilot有効/無効）を読み込み
# - GitHub API から Copilot 座席情報を取得
# - 「最終利用から7日以上経過 or シート無し」を「1週間以上未利用」とみなす
# - 以下の5列を Excel（copilot_summary_log.xlsxなど）に1行追記する：
#     データ取得日時
#     1週間以上未利用ユーザー数
#     1週間以上未利用ユーザー数(%)
#     1週間以内に利用したユーザー数
#     1週間以内に利用したユーザー数(%)
#
# 【チーム絞り込み（mall向け）】
# - .env に TARGET_TEAM_SLUG=mall を指定すると、
#   GitHub の "mall" チームメンバーに含まれるユーザーのみを対象に集計する
# - TARGET_TEAM_SLUG が未指定の場合は、user.csv に記載された全ユーザーを対象とする
#
# 【呼び出し方（PowerShell例）】
#
#   pip install requests openpyxl
#
#   python copilot_summary_append.py user.csv .env copilot_summary_log.xlsx
#
# 【.env の例】
#
#   NUXT_GITHUB_TOKEN=ghp_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
#   NUXT_PUBLIC_GITHUB_ORG=medley-inc
#   TARGET_TEAM_SLUG=mall   # 任意。指定すると mall チームメンバーのみを対象にする
#
# ============================================================

import csv
import os
import sys
from typing import Dict, Any, List, Optional, Set

import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone, timedelta

GITHUB_API_BASE = "https://api.github.com"


# =========================
# 日付パース
# =========================
def parse_iso_datetime(dt_str: str) -> Optional[datetime]:
    """GitHub API から返る ISO8601 文字列を datetime に変換する"""
    if not dt_str:
        return None
    try:
        # "Z" を "+00:00" に変換してからパース
        return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except Exception:
        return None


# =========================
# 設定ファイル(.env)読み取り
# =========================
def load_config_from_env_file(path: str = ".env") -> (str, str, Optional[str]):
    """
    .env ファイルから
      - NUXT_GITHUB_TOKEN / GITHUB_TOKEN
      - NUXT_PUBLIC_GITHUB_ORG / GITHUB_ORG
      - TARGET_TEAM_SLUG（任意）
    を読み取る
    """
    if not os.path.exists(path):
        raise RuntimeError(f"設定ファイルが見つかりません: {path}")

    env: Dict[str, str] = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or "=" not in line or line.startswith("#"):
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip()

    token = env.get("NUXT_GITHUB_TOKEN") or env.get("GITHUB_TOKEN")
    org = env.get("NUXT_PUBLIC_GITHUB_ORG") or env.get("GITHUB_ORG")
    team_slug = env.get("TARGET_TEAM_SLUG")  # 任意

    if not token:
        raise RuntimeError("設定ファイルに NUXT_GITHUB_TOKEN または GITHUB_TOKEN がありません。")
    if not org:
        raise RuntimeError("設定ファイルに NUXT_PUBLIC_GITHUB_ORG または GITHUB_ORG がありません。")

    return token, org, team_slug


def github_headers(token: str) -> Dict[str, str]:
    """GitHub API 呼び出し用ヘッダ"""
    return {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
    }


# =========================
# GitHub API：シート一覧取得
# =========================
def fetch_all_seats(org: str, token: str) -> List[Dict[str, Any]]:
    """
    /orgs/{org}/copilot/billing/seats を叩いて
    Copilot シート情報を全件取得（ページング対応）
    """
    seats: List[Dict[str, Any]] = []
    url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/billing/seats"
    headers = github_headers(token)
    params: Optional[Dict[str, Any]] = {"per_page": 100}

    while url:
        resp = requests.get(url, headers=headers, params=params)
        resp.raise_for_status()
        data = resp.json()

        seats.extend(data.get("seats", []))

        # Link ヘッダから next ページの URL を探す
        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            for part in link.split(","):
                part = part.strip()
                if 'rel="next"' in part:
                    start = part.find("<") + 1
                    end = part.find(">")
                    next_url = part[start:end]
                    break

        url = next_url
        params = None  # 2ページ目以降は URL にクエリが含まれている想定

    return seats


# =========================
# GitHub API：チームメンバー取得
# =========================
def fetch_team_members(org: str, token: str, team_slug: str) -> Set[str]:
    """
    /orgs/{org}/teams/{team_slug}/members を叩いて
    チームに所属するメンバーの login 一覧を取得する（ページング対応）
    """
    members: Set[str] = set()
    if not team_slug:
        return members

    url = f"{GITHUB_API_BASE}/orgs/{org}/teams/{team_slug}/members"
    headers = github_headers(token)
    params: Optional[Dict[str, Any]] = {"per_page": 100}

    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code == 404:
            # チームが存在しない場合など
            break
        resp.raise_for_status()
        data = resp.json()

        for m in data:
            login = m.get("login")
            if login:
                members.add(login)

        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            for part in link.split(","):
                part = part.strip()
                if 'rel="next"' in part:
                    start = part.find("<") + 1
                    end = part.find(">")
                    next_url = part[start:end]
                    break

        url = next_url
        params = None

    return members


# =========================
# CSV 読み込み
# =========================
def read_user_csv(path: str) -> List[Dict[str, str]]:
    """ユーザー一覧 CSV を読み込む"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# =========================
# メイン処理
# =========================
def main() -> None:
    if len(sys.argv) < 4:
        print("使い方: python copilot_summary_append.py user.csv .env output.xlsx", file=sys.stderr)
        sys.exit(1)

    user_csv_path = sys.argv[1]
    config_path = sys.argv[2]
    summary_path = sys.argv[3]

    # 設定ロード
    token, org, team_slug = load_config_from_env_file(config_path)

    # チームメンバー取得（TARGET_TEAM_SLUG が設定されている場合のみ）
    team_members: Optional[Set[str]] = None
    if team_slug:
        team_members = fetch_team_members(org, token, team_slug)
        print(f"TARGET_TEAM_SLUG={team_slug} が指定されました。チームメンバー {len(team_members)} 名を対象とします。")

    # ユーザー CSV 読み込み
    users = read_user_csv(user_csv_path)

    # Copilot シート情報取得
    seats = fetch_all_seats(org, token)
    seat_by_login: Dict[str, Dict[str, Any]] = {
        seat.get("assignee", {}).get("login"): seat
        for seat in seats
        if seat.get("assignee", {}).get("login")
    }

    now_utc = datetime.now(timezone.utc)

    total_users = 0
    inactive_count = 0

    # 対象ユーザーごとに「1週間以上未利用かどうか」を判定
    for u in users:
        login = u.get("GitHubユーザー名", "")

        # チーム指定がある場合は、team_members に含まれないユーザーはスキップ
        if team_members is not None and login not in team_members:
            continue

        total_users += 1

        seat = seat_by_login.get(login)
        inactive = False

        if seat:
            raw_last = seat.get("last_activity_at") or ""
            last_dt = parse_iso_datetime(raw_last)

            if last_dt is None:
                inactive = True
            else:
                delta = now_utc - last_dt.astimezone(timezone.utc)
                if delta.days >= 7:
                    inactive = True
        else:
            # シート自体が無い場合も「未利用」とみなす
            inactive = True

        if inactive:
            inactive_count += 1

    active_count = max(total_users - inactive_count, 0)

    if total_users > 0:
        inactive_pct = f"{round(inactive_count / total_users * 100)}%"
        active_pct = f"{round(active_count / total_users * 100)}%"
    else:
        inactive_pct = ""
        active_pct = ""

    # JST の現在時刻
    jst = timezone(timedelta(hours=9))
    now_jst = datetime.now(jst).strftime("%Y/%m/%d %H:%M")

    # Excel に追記
    if os.path.exists(summary_path):
        wb = load_workbook(summary_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "データ取得日時",
            "1週間以上未利用ユーザー数",
            "1週間以上未利用ユーザー数(%)",
            "1週間以内に利用したユーザー数",
            "1週間以内に利用したユーザー数(%)",
        ])

    ws.append([
        now_jst,
        inactive_count,
        inactive_pct,
        active_count,
        active_pct,
    ])

    try:
        wb.save(summary_path)
    except PermissionError as e:
        print(f"サマリーファイル '{summary_path}' を保存できませんでした。", file=sys.stderr)
        print("Excel などでファイルが開かれていないか確認し、閉じてからもう一度実行してください。", file=sys.stderr)
        print(f"詳細: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"サマリーを追記しました: {summary_path}")
    print(f"対象ユーザー数: {total_users}, 1週間以上未利用: {inactive_count}, 1週間以内利用: {active_count}")


if __name__ == "__main__":
    main()
