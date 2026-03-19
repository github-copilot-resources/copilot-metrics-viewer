#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ============================================================
# GitHub Copilot 利用状況レポート生成スクリプト
# （Copilot Last Activity Report）
#
# 【このスクリプトが行うこと】
# 1. user.csv を読み込み（各ユーザーの氏名・GitHubアカウント・参加区分など）
# 2. GitHub API を使い、Copilot の利用状況（最終利用日時・利用エディタなど）を取得
# 3. チーム指定（TARGET_TEAM_SLUG）されていれば、そのチームのメンバーだけを対象に抽出
# 4. Excel 形式でレポートを出力（./data ディレクトリに保存）
# 5. 以下を自動生成する：
#      - 対象チーム名とデータ収集日時（ヘッダ行に表示）
#      - メンバー一覧（最終利用7日以上未利用者＝赤字で強調）
#      - 未利用メンバー一覧（9項目フル）
#      - 利用/未利用ユーザー数と利用率
#
# ============================================================
# 【事前準備】
# ------------------------------------------------------------
# 1. 必要ライブラリのインストール
#       pip install requests openpyxl
#
# 2. GitHub Personal Access Token を準備（read:org 権限推奨）
#    .envファイルに以下のように記載：
#
#        NUXT_GITHUB_TOKEN=ghp_xxxxxxxxxxxxxxxxxxxxx
#        NUXT_PUBLIC_GITHUB_ORG=medley-inc
#        TARGET_TEAM_SLUG=mall        ← 任意。指定しなければ全ユーザー対象
#
# 3. user.csv の形式例（UTF-8 BOM付推奨）：
#
#        氏名,GitHubユーザー名,pmed-incへの参加,GitHub Copilotの有効/無効
#        山田 太郎,yamada,2023/04/01,有効
#        佐藤 花子,hanako,2022/11/01,無効
#
#    ※ヘッダ名称に多少揺らぎがあってもスクリプト側で自動対応します。
#
# ============================================================
# 【使い方】
# ------------------------------------------------------------
#   python copilot_last_activity_report.py user.csv
#   python copilot_last_activity_report.py user.csv .env
#
# 例：
#   python copilot_last_activity_report.py ./user.csv
#
# 実行すると、以下のように Excel ファイルが生成されます：
#
#   ./data/copilot_report_20251205113055.xlsx
#
# ============================================================
# 【出力される Excel レポートの内容】
#
#   1行目：対象チーム名（例：mall）とデータ収集日時
#   空行
#   メンバー一覧（9項目）
#     ・氏名
#     ・GitHubユーザー名
#     ・pmed-inc参加
#     ・Copilot有効/無効
#     ・シート割り当て
#     ・最終利用日時
#     ・最終利用エディタ
#     ・シート作成日
#     ・シート取消予定日
#
#   その後：
#     【1週間以上未利用ユーザー（赤字対象）一覧】
#     ・未利用ユーザー一覧（同じ9項目）
#
#     利用/未利用ユーザー数・割合（%）
#
# ============================================================
# 【注意事項】
# ------------------------------------------------------------
# - GitHub API 呼び出しが多いので、レートリミットに注意
# - Excel を開いた状態では保存に失敗するので閉じて再実行
# - user.csv のヘッダ揺らぎ（「pmed-inc参加」「pmed-incへの参加」など）は
#   自動的に正規化して解決します
#
# ============================================================
# 【キャッシュとETagについて】
# ------------------------------------------------------------
# このスクリプトは GitHub API の呼び出し回数を削減するため、
# キャッシュ機構を実装しています。
#
# ■ キャッシュの種類と動作
#
# 1. ETagベースのキャッシュ（Copilotシート情報、チームメンバー）
#    - GitHub API の ETag ヘッダーを利用した変更検知
#    - If-None-Match ヘッダーで前回の ETag を送信
#    - データ未変更時: 304 Not Modified レスポンス（軽量、キャッシュ使用）
#    - データ変更時: 200 OK レスポンス（新データ取得、キャッシュ更新）
#    - メリット: データが変更された瞬間に検知、時間制限なし
#
# 2. 時間ベースのキャッシュ（ユーザー詳細情報）
#    - 有効期限: 60分（CACHE_EXPIRY_MINUTES_USER_DETAILS で設定可能）
#    - 60分以内の再実行時はキャッシュから読み込み（API呼び出しなし）
#    - 60分経過後は API から再取得
#    - メリット: エディタ情報も含めて大幅にAPI呼び出しを削減
#
# ■ キャッシュファイルの保存場所
#    ./cache ディレクトリ（自動作成）
#    - JSON形式で保存
#    - ファイル名: MD5ハッシュ（org名、チーム名、ユーザー名などから生成）
#
# ■ API呼び出し削減効果（48ユーザー、チーム指定の場合）
#    初回実行: 約50回のAPI呼び出し
#    60分以内の再実行: 約2回（シート + チーム、両方とも304レスポンス）
#    削減率: 96%
#
# ■ キャッシュの無効化
#    ./cache ディレクトリを削除すると、全てのキャッシュがクリアされます
#
# ============================================================
# 【ChangeLog】
# 2025-12-08: Copilot使用状況メトリクス追加と組織全体APIへの移行
#   - 4つの使用状況メトリクス列を追加 (計13列に拡張)
#     * 総提案数 (total_suggestions_count): 過去28日間のコード提案総数
#     * 総承認数 (total_acceptances_count): 過去28日間の承認総数
#     * Chat総ターン数 (total_chat_turns): Copilot Chatでの対話ターン数
#     * Chat承認数 (total_chat_acceptances): Chatでのコード挿入承認数
#   - メトリクス取得方式の改善 (組織全体API活用)
#     * 旧方式: ユーザーごとに個別API呼び出し (57ユーザー = 57回以上のAPI呼び出し)
#     * 新方式: 組織全体メトリクスを1回で取得し、ユーザー別に分解
#     * API呼び出し削減: 57+ 回 → 1回 (98%以上削減、パフォーマンス大幅向上)
#   - 時間ベースキャッシュの導入 (組織メトリクスとユーザー詳細)
#     * 60分間キャッシュを保持し、頻繁な再実行時にAPI負荷を軽減
#     * ETagベースキャッシュと併用して最適化
#   - デバッグログ強化
#     * メトリクス取得の詳細ログ (日数、ユーザー数、トップユーザーなど)
#     * キャッシュヒット/ミスの可視化
# 
# 2025-12-05: コード品質改善とパフォーマンス最適化
#   - マジックナンバーの定数化 (INACTIVE_DAYS=7, MAX_WORKERS=物理コア数-1)
#   - ログ機能追加 (logging モジュール、処理進捗の可視化)
#   - ページネーション処理の共通化 (paginate_github_api 関数)
#   - エラーハンドリング強化 (timeout=30秒、具体的な例外型、lazy formatting)
#   - ユーザー詳細取得の並列化 (ThreadPoolExecutor、物理コア数-1並列、最小1)
#     * 48ユーザーの処理が約24秒 → 約3.7秒に短縮（約6.5倍高速化）
#     * マルチコアCPUを最適活用（システムに1コア分の余裕を残す）
#     * 8コアCPUなら7並列、16コアなら15並列で自動調整
#   - タイムゾーン処理の統一 (UTC基準)
#   - CSV ヘッダ検出の改善 (完全一致を優先)
#   - ETagベースのキャッシュ機構実装
#     * GitHub API の ETag を利用した変更検知
#     * 304 Not Modified レスポンスによる API 呼び出し削減
#     * ./cache ディレクトリにキャッシュデータを保存
#     * データ変更時のみ新規取得、未変更時はキャッシュ使用
# ============================================================

import csv
import hashlib
import json
import logging
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed  # type: ignore
from typing import Dict, Any, List, Optional, Set

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timezone, timedelta

GITHUB_API_BASE = "https://api.github.com"
GITHUB_API_VERSION = "2026-03-10"
INACTIVE_DAYS = 7  # 未利用と判定する日数
# 並列API呼び出しの最大スレッド数（物理コア数-1、最小1）
MAX_WORKERS = max(1, os.cpu_count() - 1) if os.cpu_count() else 10
JST_OFFSET = 9  # JSTのオフセット（UTC+9）

# キャッシュ設定
CACHE_DIR = "./cache"
CACHE_EXPIRY_MINUTES = 10  # キャッシュ有効期限（分）- ETag対応 API用（参考情報）
CACHE_EXPIRY_MINUTES_USER_DETAILS = 60  # ユーザー詳細情報のキャッシュ有効期限（分）

# ログ設定
LOG_DIR = "./logs"
os.makedirs(LOG_DIR, exist_ok=True)

# ログファイル名（タイムスタンプ付き）
log_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file = os.path.join(LOG_DIR, f"debug_{log_timestamp}.log")

# ロガーの設定
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

# フォーマッター
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# コンソールハンドラ（標準出力）
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)  # コンソールはINFO以上
console_handler.setFormatter(formatter)

# ファイルハンドラ（デバッグログファイル）
file_handler = logging.FileHandler(log_file, encoding='utf-8')
file_handler.setLevel(logging.DEBUG)  # ファイルは全てのログを記録
file_handler.setFormatter(formatter)

# ハンドラを追加
logger.addHandler(console_handler)
logger.addHandler(file_handler)

logging.info("=" * 60)
logging.info("GitHub Copilot 利用状況レポート生成スクリプト 開始")
logging.info("デバッグログ: %s", log_file)
logging.info("MAX_WORKERS: %d (物理コア数: %d)", MAX_WORKERS, os.cpu_count() or 0)
logging.info("=" * 60)


# -------------------------
# キャッシュユーティリティ
# -------------------------
def get_cache_key(prefix: str, *args: Any) -> str:
    """キャッシュキーを生成"""
    key_str = f"{prefix}_{'_'.join(str(arg) for arg in args)}"
    return hashlib.md5(key_str.encode()).hexdigest()


def load_cache(cache_key: str) -> tuple[Optional[Any], Optional[str]]:
    """キャッシュからデータとETagを読み込む"""
    if not os.path.exists(CACHE_DIR):
        return None, None
    
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    if not os.path.exists(cache_file):
        return None, None
    
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
        
        cached_at = cache_data.get('cached_at', 0)
        etag = cache_data.get('etag')
        data = cache_data.get('data')
        
        now = time.time()
        age_minutes = (now - cached_at) / 60
        
        logging.info("キャッシュが存在: %s (取得から%.1f分経過, ETag: %s)", 
                     cache_key, age_minutes, etag or "N/A")
        
        return data, etag
    
    except (json.JSONDecodeError, IOError) as e:
        logging.warning("キャッシュ読み込みエラー: %s - %s", cache_key, e)
        return None, None


def load_cache_with_expiry(cache_key: str, expiry_minutes: int) -> Optional[Any]:
    """時間ベースのキャッシュ読み込み（有効期限チェック付き）"""
    if not os.path.exists(CACHE_DIR):
        return None
    
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    if not os.path.exists(cache_file):
        return None
    
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
        
        cached_at = cache_data.get('cached_at', 0)
        now = time.time()
        age_minutes = (now - cached_at) / 60
        
        # 有効期限チェック
        if age_minutes > expiry_minutes:
            logging.debug("キャッシュが期限切れ: %s (%.1f分経過)", cache_key, age_minutes)
            return None
        
        logging.info("キャッシュからデータを読み込み: %s (%.1f分経過)", cache_key, age_minutes)
        return cache_data.get('data')
    
    except (json.JSONDecodeError, IOError) as e:
        logging.warning("キャッシュ読み込みエラー: %s - %s", cache_key, e)
        return None


def save_cache(cache_key: str, data: Any, etag: Optional[str] = None) -> None:
    """データとETagをキャッシュに保存"""
    os.makedirs(CACHE_DIR, exist_ok=True)
    
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    cache_data = {
        'cached_at': time.time(),
        'etag': etag,
        'data': data
    }
    
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        logging.debug("キャッシュに保存: %s (ETag: %s)", cache_key, etag or "N/A")
    except IOError as e:
        logging.warning("キャッシュ保存エラー: %s - %s", cache_key, e)


# -------------------------
# 日付パース & フォーマット
# -------------------------
def parse_iso_datetime(dt_str: str) -> Optional[datetime]:
    """GitHub API から返る ISO8601 文字列を datetime に変換する"""
    if not dt_str:
        return None
    try:
        # "Z" を "+00:00" に変換してからパース
        return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return None


def format_datetime(dt_str: str) -> str:
    """ISO8601 文字列を 'yyyy/MM/dd HH:MM' 形式 (JST) に変換する"""
    dt = parse_iso_datetime(dt_str)
    if not dt:
        return ""
    # UTCからJSTに変換
    jst = timezone(timedelta(hours=JST_OFFSET))
    dt_jst = dt.astimezone(jst)
    return dt_jst.strftime("%Y/%m/%d %H:%M")


# -------------------------
# 設定ファイル(.env)読み取り
# -------------------------
def load_config_from_env_file(path: str = ".env") -> tuple[str, str, Optional[str]]:
    token = None
    org = None
    team_slug: Optional[str] = None

    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if k in ("NUXT_GITHUB_TOKEN", "GITHUB_TOKEN"):
                    token = v
                elif k in ("NUXT_PUBLIC_GITHUB_ORG", "GITHUB_ORG"):
                    org = v
                elif k == "TARGET_TEAM_SLUG":
                    team_slug = v

    token = os.environ.get("NUXT_GITHUB_TOKEN") or os.environ.get("GITHUB_TOKEN") or token
    org = os.environ.get("NUXT_PUBLIC_GITHUB_ORG") or os.environ.get("GITHUB_ORG") or org
    team_slug = os.environ.get("TARGET_TEAM_SLUG") or team_slug

    if not token:
        print("GitHub Token が見つかりません。.env または環境変数を確認してください。", file=sys.stderr)
        sys.exit(1)
    if not org:
        print("GitHub Organization が見つかりません。.env または環境変数を確認してください。", file=sys.stderr)
        sys.exit(1)

    return token, org, team_slug


# -------------------------
# GitHub API 共通ヘッダ
# -------------------------
def github_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": GITHUB_API_VERSION,
    }


def download_usage_metrics_records(download_links: List[str]) -> List[Dict[str, Any]]:
    """署名付きURLからUsage Metricsレポート本体をダウンロードして展開する"""
    records: List[Dict[str, Any]] = []

    for download_url in download_links:
        resp = requests.get(download_url, timeout=60)
        resp.raise_for_status()

        try:
            payload = resp.json()
        except json.JSONDecodeError:
            payload = None

        if isinstance(payload, list):
            records.extend(item for item in payload if isinstance(item, dict))
            continue

        if isinstance(payload, dict):
            records.append(payload)
            continue

        for line in resp.text.strip().splitlines():
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(record, dict):
                records.append(record)

    return records


def fetch_usage_metrics_report_records(url: str, token: str, cache_key: str, description: str) -> tuple[List[Dict[str, Any]], Optional[str]]:
    """Copilot Usage Metrics APIのmanifestを取得し、download_links配下のレコードを返す"""
    cached_data = load_cache_with_expiry(cache_key, CACHE_EXPIRY_MINUTES_USER_DETAILS)
    if cached_data is not None:
        logging.info("%s をキャッシュから読み込み", description)
        return cached_data, None

    logging.info("%s を取得中: %s", description, url)
    resp = requests.get(url, headers=github_headers(token), timeout=30)

    if resp.status_code == 403:
        logging.warning("%s へのアクセスが拒否されました", description)
        logging.warning("Copilot usage metrics policy とトークン権限を確認してください")
        return [], "forbidden"

    if resp.status_code == 404:
        logging.warning("%s が見つかりません", description)
        logging.warning("組織設定、権限、またはAPIサポート状況を確認してください")
        return [], "not_found"

    resp.raise_for_status()
    manifest = resp.json()
    download_links = manifest.get("download_links", []) if isinstance(manifest, dict) else []
    if not download_links:
        logging.warning("%s の download_links が空です", description)
        return [], "empty"

    records = download_usage_metrics_records(download_links)
    save_cache(cache_key, records)
    logging.info("%s から %d レコード取得", description, len(records))
    return records, None


def normalize_org_usage_metrics_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Usage Metricsの組織レポートを、日次行ベースの扱いやすい形式に正規化する"""
    normalized: List[Dict[str, Any]] = []

    for record in records:
        day_totals = record.get("day_totals")
        if not isinstance(day_totals, list):
            continue

        for day in day_totals:
            if not isinstance(day, dict):
                continue

            pull_requests = day.get("pull_requests", {}) if isinstance(day.get("pull_requests"), dict) else {}
            normalized.append(
                {
                    "_source": "usage_metrics",
                    "date": day.get("day", ""),
                    "daily_active_users": day.get("daily_active_users", 0),
                    "weekly_active_users": day.get("weekly_active_users", 0),
                    "monthly_active_users": day.get("monthly_active_users", 0),
                    "daily_active_cli_users": day.get("daily_active_cli_users", 0),
                    "user_initiated_interaction_count": day.get("user_initiated_interaction_count", 0),
                    "code_generation_activity_count": day.get("code_generation_activity_count", 0),
                    "code_acceptance_activity_count": day.get("code_acceptance_activity_count", 0),
                    "loc_suggested_to_add_sum": day.get("loc_suggested_to_add_sum", 0),
                    "loc_added_sum": day.get("loc_added_sum", 0),
                    "pull_requests_total_created": pull_requests.get("total_created", 0),
                    "pull_requests_total_merged": pull_requests.get("total_merged", 0),
                }
            )

    normalized.sort(key=lambda item: item.get("date", ""))
    return normalized


# -------------------------
# GitHub API ページネーション共通処理
# -------------------------
def paginate_github_api(url: str, token: str, data_key: Optional[str] = None, etag: Optional[str] = None) -> tuple[List[Dict[str, Any]], Optional[str]]:
    """
    GitHub API のページネーションを処理して全データを取得 (ETag対応)
    
    Args:
        url: API エンドポイント URL
        token: GitHub トークン
        data_key: レスポンスからデータを取得するキー（Noneの場合はレスポンス自体がリスト）
        etag: 前回取得時のETag
    
    Returns:
        (データリスト, 新しいETag)
    """
    results: List[Dict[str, Any]] = []
    params: Optional[Dict[str, Any]] = {"per_page": 100}
    new_etag: Optional[str] = None
    
    while url:
        try:
            headers = github_headers(token)
            # ETagがあればIf-None-Matchヘッダを追加
            if etag and not results:  # 最初のリクエストのみ
                headers['If-None-Match'] = etag
            
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            
            # 304 Not Modified - データ未変更
            if resp.status_code == 304:
                logging.info("データ未変更 (304 Not Modified) - キャッシュを使用")
                return [], etag  # 空リストと旧ETagを返す（呼び出し側でキャッシュ使用）
            
            resp.raise_for_status()
            
            # 新しいETagを取得（最初のレスポンスのみ）
            if not new_etag:
                new_etag = resp.headers.get('ETag')
            
            data = resp.json()
            
            if data_key:
                results.extend(data.get(data_key, []))
            else:
                if isinstance(data, list):
                    results.extend(data)
                else:
                    results.append(data)
            
            # 次ページのURLを取得
            link = resp.headers.get("Link", "")
            next_url = None
            if link:
                for part in link.split(","):
                    part = part.strip()
                    if 'rel="next"' in part:
                        start = part.find("<") + 1
                        end = part.find(">")
                        next_url = part[start:end]
                        break
            
            url = next_url
            params = None
            
        except requests.exceptions.RequestException as e:
            logging.error("API リクエストエラー: %s - %s", url, e)
            break
    
    return results, new_etag


# -------------------------
# Copilot シート情報取得
# -------------------------
def fetch_all_seats(org: str, token: str) -> List[Dict[str, Any]]:
    """Copilot シート情報を全て取得（ETagキャッシュ対応）"""
    cache_key = get_cache_key("seats", org)
    
    # キャッシュチェック
    cached_data, cached_etag = load_cache(cache_key)
    
    # API呼び出し
    url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/billing/seats"
    logging.info("Copilot シート情報を取得中: %s (ETag付きリクエスト)", org)
    
    seats, new_etag = paginate_github_api(url, token, data_key="seats", etag=cached_etag)
    
    # 304の場合はキャッシュを使用
    if not seats and cached_data is not None:
        logging.info("キャッシュデータを使用: %d件", len(cached_data))
        return cached_data
    
    # 新しいデータをキャッシュに保存
    if seats:
        logging.info("新しいデータを取得: %d件", len(seats))
        save_cache(cache_key, seats, new_etag)
        return seats
    
    # キャッシュも新データもない場合
    return []


# -------------------------
# チームメンバー一覧取得
# -------------------------
def fetch_team_members(org: str, token: str, team_slug: str) -> Set[str]:
    """チームメンバー一覧を取得（ETagキャッシュ対応）"""
    cache_key = get_cache_key("team_members", org, team_slug)
    
    # キャッシュチェック
    cached_data, cached_etag = load_cache(cache_key)
    
    url = f"{GITHUB_API_BASE}/orgs/{org}/teams/{team_slug}/members"
    logging.info("チームメンバーを取得中: %s (ETag付きリクエスト)", team_slug)
    
    try:
        # 404チェックのため最初のリクエストを個別に実行
        headers = github_headers(token)
        if cached_etag:
            headers['If-None-Match'] = cached_etag
        
        resp = requests.get(url, headers=headers, params={"per_page": 100}, timeout=30)
        
        if resp.status_code == 404:
            logging.error("指定されたチームが見つかりません: org=%s, team_slug=%s", org, team_slug)
            return set()
        
        # 304 Not Modified - キャッシュ使用
        if resp.status_code == 304:
            logging.info("データ未変更 (304) - キャッシュを使用: %d名", len(cached_data) if cached_data else 0)
            return set(cached_data) if cached_data else set()
        
        resp.raise_for_status()
        
        # ページネーション処理
        data, new_etag = paginate_github_api(url, token, etag=cached_etag)
        members = {m.get("login") for m in data if m.get("login")}
        logging.info("チームメンバー %d 名を取得しました", len(members))
        
        # キャッシュ保存（setはJSON化できないのlistに変換）
        save_cache(cache_key, list(members), new_etag)
        
        return members
        
    except requests.exceptions.RequestException as e:
        logging.error("チームメンバー取得エラー: %s", e)
        return set()


# -------------------------
# 個別ユーザー Copilot 詳細
# -------------------------
def fetch_user_seat_details(org: str, token: str, username: str) -> Optional[Dict[str, Any]]:
    """個別ユーザーの Copilot 詳細情報を取得（キャッシュ対応）"""
    cache_key = get_cache_key("user_details", org, username)
    
    # キャッシュチェック
    cached_data = load_cache_with_expiry(cache_key, CACHE_EXPIRY_MINUTES_USER_DETAILS)
    if cached_data is not None:
        return cached_data
    
    url = f"{GITHUB_API_BASE}/orgs/{org}/members/{username}/copilot"
    try:
        resp = requests.get(url, headers=github_headers(token), timeout=30)
        
        if resp.status_code in (404, 422):
            return None
        
        resp.raise_for_status()
        details = resp.json()
        
        # キャッシュ保存
        save_cache(cache_key, details)
        
        return details
        
    except requests.exceptions.RequestException as e:
        logging.warning("ユーザー %s の詳細取得エラー: %s", username, e)
        return None


# -------------------------
# 組織全体のメトリクス取得（全ユーザー分）
# -------------------------
def fetch_org_usage_metrics(org: str, token: str, since: str, until: str) -> List[Dict[str, Any]]:
    """
    組織全体のCopilot使用統計を取得（新Usage Metrics APIを優先、不可ならlegacyへフォールバック）
    
    Args:
        org: 組織名
        token: GitHubトークン
        since: 開始日（YYYY-MM-DD）
        until: 終了日（YYYY-MM-DD）
    
    Returns:
        日別の組織メトリクスデータのリスト
    """
    cache_key = get_cache_key("org_usage_metrics_v2", org, since, until)
    
    # キャッシュチェック
    cached_data = load_cache_with_expiry(cache_key, CACHE_EXPIRY_MINUTES_USER_DETAILS)
    if cached_data is not None:
        logging.info("組織メトリクスをキャッシュから読み込み")
        return cached_data
    
    try:
        report_url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/metrics/reports/organization-28-day/latest"
        report_records, _ = fetch_usage_metrics_report_records(
            report_url,
            token,
            cache_key,
            "Organization Usage Metrics API"
        )
        data = normalize_org_usage_metrics_records(report_records)

        if data:
            total_active = sum(day.get("daily_active_users", 0) for day in data)
            total_interactions = sum(day.get("user_initiated_interaction_count", 0) for day in data)
            logging.info("Usage Metrics API: %d 日分のデータを取得", len(data))
            logging.info("期間中の延べ日次アクティブユーザー数: %d", total_active)
            logging.info("期間中のユーザー起点インタラクション数: %d", total_interactions)
            save_cache(cache_key, data)
            return data

        logging.warning("新しいOrganization Usage Metrics APIから日次データを取得できませんでした。legacy APIにフォールバックします")

        legacy_url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/metrics"
        params = {
            "since": since,
            "until": until,
            "per_page": 100
        }
        resp = requests.get(legacy_url, headers=github_headers(token), params=params, timeout=30)

        if resp.status_code == 404:
            logging.error("legacy Copilot Metrics APIが見つかりません")
            return []

        if resp.status_code == 403:
            logging.error("legacy Copilot Metrics APIへのアクセスが拒否されました")
            return []

        if resp.status_code == 422:
            logging.error("Copilot metrics policy または legacy API 設定が無効です")
            return []

        resp.raise_for_status()
        legacy_data = resp.json()
        for day in legacy_data:
            if isinstance(day, dict):
                day["_source"] = "legacy_metrics"

        save_cache(cache_key, legacy_data)
        return legacy_data
        
    except requests.exceptions.RequestException as e:
        logging.error("組織メトリクス取得エラー: %s", e)
        return []


# -------------------------
# Copilot Usage Metrics API（NDJSONレポート）
# -------------------------
def fetch_usage_metrics_report(org: str, token: str, days: int = 28) -> tuple[Dict[str, Dict[str, Any]], Optional[str]]:
    """
    Copilot Usage Metrics APIから組織単位のユーザー別詳細データを取得
    
    Args:
        org: 組織名
        token: GitHubトークン（read:org または Organization Copilot metrics 権限が必要）
        days: 取得日数（デフォルト28日）
    
    Returns:
        ユーザー名をキーとした辞書。値は各ユーザーの集計データ:
        {
            "user_login": {
                "total_interactions": int,  # user_initiated_interaction_count の合計
                "total_generations": int,   # code_generation_activity_count の合計
                "total_acceptances": int,   # code_acceptance_activity_count の合計
                "total_loc_added": int,     # loc_added_sum の合計
                "total_loc_deleted": int,   # loc_deleted_sum の合計
                "total_loc_suggested_deleted": int,
                "days_active": int,         # アクティブな日数
                "days_used_chat": int,
                "days_used_cli": int,
                "days_used_agent": int,
                "cli_prompt_count": int,
                "cli_request_count": int,
                "cli_session_count": int,
                "ides": {},                 # IDE別内訳
                "languages": {},            # 言語別内訳
                "features": {}              # 機能別内訳
            }
        }
    """
    cache_key = get_cache_key("user_usage_metrics_v2", org, days)
    
    # キャッシュチェック（60分有効期限）
    cached_data = load_cache_with_expiry(cache_key, CACHE_EXPIRY_MINUTES_USER_DETAILS)
    if cached_data is not None:
        logging.info("Usage Metricsをキャッシュから読み込み")
        return cached_data, None
    
    if days != 28:
        logging.warning("最新APIではこの実装は28日レポートを前提にしています。days=%d は28日に丸めます", days)

    url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/metrics/reports/users-28-day/latest"
    
    try:
        logging.info("Organization Users Usage Metrics API呼び出し中: %s (直近28日分)", org)
        report_records, fetch_status = fetch_usage_metrics_report_records(
            url,
            token,
            cache_key,
            "Organization Users Usage Metrics API"
        )

        if not report_records:
            if fetch_status == "forbidden":
                return {}, "ユーザー別 Usage Metrics は 403 のため未取得です"
            if fetch_status == "not_found":
                return {}, "ユーザー別 Usage Metrics API が見つかりません"
            if fetch_status == "empty":
                return {}, "ユーザー別 Usage Metrics に取得可能なデータがありません"
            return {}, "ユーザー別 Usage Metrics を取得できませんでした"

        user_metrics: Dict[str, Dict[str, Any]] = {}

        for record in report_records:
            try:
                user_login = record.get("user_login")
                if not user_login:
                    continue
                
                # ユーザーデータ初期化
                if user_login not in user_metrics:
                    user_metrics[user_login] = {
                        "total_interactions": 0,
                        "total_generations": 0,
                        "total_acceptances": 0,
                        "total_loc_added": 0,
                        "total_loc_suggested": 0,
                        "total_loc_deleted": 0,
                        "total_loc_suggested_deleted": 0,
                        "days_active": 0,
                        "days_used_chat": 0,
                        "days_used_cli": 0,
                        "days_used_agent": 0,
                        "cli_prompt_count": 0,
                        "cli_request_count": 0,
                        "cli_session_count": 0,
                        "ides": {},
                        "languages": {},
                        "features": {}
                    }
                
                # 集計
                user_metrics[user_login]["total_interactions"] += record.get("user_initiated_interaction_count", 0)
                user_metrics[user_login]["total_generations"] += record.get("code_generation_activity_count", 0)
                user_metrics[user_login]["total_acceptances"] += record.get("code_acceptance_activity_count", 0)
                user_metrics[user_login]["total_loc_added"] += record.get("loc_added_sum", 0)
                user_metrics[user_login]["total_loc_suggested"] += record.get("loc_suggested_to_add_sum", 0)
                user_metrics[user_login]["total_loc_deleted"] += record.get("loc_deleted_sum", 0)
                user_metrics[user_login]["total_loc_suggested_deleted"] += record.get("loc_suggested_to_delete_sum", 0)
                
                # アクティブ日数（interaction > 0 の日）
                if record.get("user_initiated_interaction_count", 0) > 0:
                    user_metrics[user_login]["days_active"] += 1
                if record.get("used_chat"):
                    user_metrics[user_login]["days_used_chat"] += 1
                if record.get("used_cli"):
                    user_metrics[user_login]["days_used_cli"] += 1
                if record.get("used_agent"):
                    user_metrics[user_login]["days_used_agent"] += 1

                cli_totals = record.get("totals_by_cli", {}) if isinstance(record.get("totals_by_cli"), dict) else {}
                user_metrics[user_login]["cli_prompt_count"] += cli_totals.get("prompt_count", 0)
                user_metrics[user_login]["cli_request_count"] += cli_totals.get("request_count", 0)
                user_metrics[user_login]["cli_session_count"] += cli_totals.get("session_count", 0)
                
                # IDE別内訳
                for ide_data in record.get("totals_by_ide", []):
                    ide_name = ide_data.get("ide") or ide_data.get("name") or "unknown"
                    ide_score = (
                        ide_data.get("user_initiated_interaction_count", 0)
                        + ide_data.get("code_generation_activity_count", 0)
                        + ide_data.get("code_acceptance_activity_count", 0)
                        + ide_data.get("loc_added_sum", 0)
                    )
                    if ide_name not in user_metrics[user_login]["ides"]:
                        user_metrics[user_login]["ides"][ide_name] = 0
                    user_metrics[user_login]["ides"][ide_name] += ide_score
                
                # 言語別内訳
                for lang_data in record.get("totals_by_language_feature", []):
                    lang_name = lang_data.get("language") or lang_data.get("name") or "unknown"
                    lang_score = (
                        lang_data.get("code_generation_activity_count", 0)
                        + lang_data.get("code_acceptance_activity_count", 0)
                        + lang_data.get("loc_added_sum", 0)
                    )
                    if lang_name not in user_metrics[user_login]["languages"]:
                        user_metrics[user_login]["languages"][lang_name] = 0
                    user_metrics[user_login]["languages"][lang_name] += lang_score

                # 機能別内訳
                for feature_data in record.get("totals_by_feature", []):
                    feature_name = feature_data.get("feature") or feature_data.get("name") or "unknown"
                    feature_score = (
                        feature_data.get("user_initiated_interaction_count", 0)
                        + feature_data.get("code_generation_activity_count", 0)
                        + feature_data.get("code_acceptance_activity_count", 0)
                        + feature_data.get("loc_added_sum", 0)
                    )
                    if feature_name not in user_metrics[user_login]["features"]:
                        user_metrics[user_login]["features"][feature_name] = 0
                    user_metrics[user_login]["features"][feature_name] += feature_score
                    
            except (KeyError, TypeError) as e:
                logging.debug("ユーザーレコードのパースエラー: %s", e)
                continue
        
        logging.info("Usage Metrics: %d ユーザー分のデータを取得", len(user_metrics))
        
        # トップ5ユーザーをログ出力
        sorted_users = sorted(user_metrics.items(), 
                            key=lambda x: x[1]["total_acceptances"], 
                            reverse=True)[:5]
        for i, (username, data) in enumerate(sorted_users, 1):
            logging.info("  %d. %s: %d acceptances, %d LoC added", 
                        i, username, data["total_acceptances"], data["total_loc_added"])
        
        # キャッシュに保存
        save_cache(cache_key, user_metrics)
        
        return user_metrics, None
        
    except requests.exceptions.RequestException as e:
        logging.error("Usage Metrics取得エラー: %s", e)
        logging.info("Usage Metrics APIが利用できません。基本情報のみでレポートを生成します")
        return {}, "ユーザー別 Usage Metrics の取得中にエラーが発生しました"


# -------------------------
# CSV 読み込み（ヘッダ正規化 & N/A 補完）
# -------------------------
def read_user_csv(path: str) -> List[Dict[str, str]]:
    """
    ユーザー一覧 CSV を読み込む。
    - ヘッダ名の前後の空白、全角空白、複数空白を正規化
    - 各値も同様に正規化
    - 空値は "N/A" を入れる
    """
    import re

    def normalize(s: Any) -> Any:
        if not isinstance(s, str):
            return s
        s = s.replace("\u3000", " ")
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    rows: List[Dict[str, str]] = []

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)

        if reader.fieldnames:
            reader.fieldnames = [normalize(h) for h in reader.fieldnames]

        for raw_row in reader:
            normalized_row: Dict[str, str] = {}
            for k, v in raw_row.items():
                key = normalize(k)
                val = normalize(v)
                if val in ("", None):
                    val = "N/A"
                normalized_row[key] = val
            rows.append(normalized_row)

    return rows


# -------------------------
# メイン処理
# -------------------------
def main() -> None:
    if len(sys.argv) < 2:
        print("使い方: python copilot_last_activity_report.py user.csv [.env]", file=sys.stderr)
        sys.exit(1)

    user_csv_path = sys.argv[1]
    config_path = sys.argv[2] if len(sys.argv) >= 3 else ".env"

    token, org, team_slug = load_config_from_env_file(config_path)

    team_members: Optional[Set[str]] = None
    if team_slug:
        team_members = fetch_team_members(org, token, team_slug)
        print(f"TARGET_TEAM_SLUG={team_slug} が指定されました。チームメンバー {len(team_members)} 名を対象とします。")

    seats = fetch_all_seats(org, token)
    seat_by_login: Dict[str, Dict[str, Any]] = {
        seat.get("assignee", {}).get("login"): seat
        for seat in seats
        if seat.get("assignee", {}).get("login")
    }

    users = read_user_csv(user_csv_path)
    if not users:
        print("user.csv にデータがありません。", file=sys.stderr)
        sys.exit(1)

    # === ここで実際のヘッダ名から列名を推定する ===
    sample_keys = list(users[0].keys())

    def find_col(patterns: List[str], default: str) -> str:
        """CSVヘッダから列名を検出（完全一致を優先）"""
        # まず完全一致を試す
        if default in sample_keys:
            return default
        # 次にパターンマッチ
        for col in sample_keys:
            if all(p in col for p in patterns):
                return col
        return default

    # あなたの user.csv だと:
    #  - 「pmed-incへの参加」
    #  - 「GitHub Copilotの有効/無効」
    joined_col = find_col(["参加"], "pmed-incへの参加")
    enabled_col = find_col(["Copilot", "有効"], "GitHub Copilotの有効/無効")

    wb = Workbook()
    ws = wb.active
    ws.title = "copilot_report"

    team_label = team_slug if team_slug else "全体"
    now_utc = datetime.now(timezone.utc)
    # JSTで表示
    jst = timezone(timedelta(hours=JST_OFFSET))
    now_jst = now_utc.astimezone(jst)
    collected_at_str = now_jst.strftime("%Y/%m/%d %H:%M")
    ws.append([f"対象チーム: {team_label}", f"データ収集日時: {collected_at_str}"])
    ws.append([])

    header = [
        "氏名",
        "GitHubユーザー名",
        "pmed-inc参加",
        "Copilot有効/無効",
        "シート割り当て",
        "最終利用日時",
        "最終利用エディタ",
        "シート作成日",
        "シート取消予定日",
        "プロンプト数(28日)",
        "コード生成数(28日)",
        "受け入れ数(28日)",
        "追加LoC(28日)",
        "アクティブ日数(28日)",
        "主要IDE",
    ]
    ws.append(header)

    inactive_users: List[Dict[str, str]] = []
    reported_users = 0
    
    logging.info("ユーザー情報処理開始: %d 名", len(users))

    def nz(v: Any) -> str:
        if v in ("", None):
            return "N/A"
        return str(v)

    # フィルタリング済みユーザーリストを作成
    target_users = []
    for user in users:
        login = nz(user.get("GitHubユーザー名"))
        if login == "N/A":
            continue
        if team_members is not None and login not in team_members:
            continue
        target_users.append(user)
    
    logging.info("対象ユーザー: %d 名", len(target_users))
    
    # メトリクス取得の日付範囲を設定（過去28日間）
    until_date = now_jst.strftime("%Y-%m-%d")
    since_date = (now_jst - timedelta(days=28)).strftime("%Y-%m-%d")
    
    # 組織全体のメトリクスを一度に取得（統計シート用、キャッシュに保存）
    logging.info("組織全体のメトリクスを取得中...")
    org_metrics_data = fetch_org_usage_metrics(org, token, since_date, until_date)
    
    # Usage Metrics API（NDJSON）でユーザー別詳細データを取得
    logging.info("Usage Metrics API（ユーザー別詳細）を取得中...")
    usage_metrics_data, usage_metrics_warning = fetch_usage_metrics_report(org, token, days=28)

    if usage_metrics_warning:
        ws.append([f"Usage Metrics(28日): {usage_metrics_warning}"])
        ws.append([])
    
    # シートを持つユーザーの詳細情報を並列取得
    logins_with_seat = [u.get("GitHubユーザー名") for u in target_users 
                        if seat_by_login.get(u.get("GitHubユーザー名"))]
    
    user_details: Dict[str, Optional[Dict[str, Any]]] = {}
    
    if logins_with_seat:
        logging.info("ユーザー詳細情報を並列取得中: %d 名", len(logins_with_seat))
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # 詳細情報取得
            future_to_login_details = {
                executor.submit(fetch_user_seat_details, org, token, login): login
                for login in logins_with_seat
            }
            
            # 詳細情報を収集
            for future in as_completed(future_to_login_details):
                login = future_to_login_details[future]
                try:
                    user_details[login] = future.result()
                except (requests.exceptions.RequestException, ValueError, KeyError) as e:
                    logging.error("ユーザー %s の詳細取得で例外発生: %s", login, e)
                    user_details[login] = None
    
    # Excel書き込み
    for user in target_users:
        name = nz(user.get("氏名"))
        login = nz(user.get("GitHubユーザー名"))
        joined = nz(user.get(joined_col))
        enabled = nz(user.get(enabled_col))

        seat = seat_by_login.get(login)
        has_seat = "あり" if seat else "なし"

        inactive = False
        last_disp = ""
        created_disp = ""
        cancel_disp = ""
        editor = ""

        if seat:
            raw_last = seat.get("last_activity_at") or ""
            raw_created = seat.get("created_at") or ""
            raw_cancel = seat.get("pending_cancellation_date") or ""

            last_disp = format_datetime(raw_last)
            created_disp = format_datetime(raw_created)
            cancel_disp = format_datetime(raw_cancel)

            details = user_details.get(login)
            editor = details.get("last_activity_editor") if details else ""

            last_dt = parse_iso_datetime(raw_last)
            if last_dt is None:
                inactive = True
            else:
                delta = now_utc - last_dt.astimezone(timezone.utc)
                if delta.days >= INACTIVE_DAYS:
                    inactive = True
        else:
            inactive = True
        
        # Usage Metricsデータを取得
        user_usage = usage_metrics_data.get(login, {})
        if usage_metrics_warning:
            total_interactions = ""
            total_generations = ""
            total_acceptances = ""
            total_loc_added = ""
            days_active = ""
        else:
            total_interactions = user_usage.get("total_interactions", 0) if user_usage else 0
            total_generations = user_usage.get("total_generations", 0) if user_usage else 0
            total_acceptances = user_usage.get("total_acceptances", 0) if user_usage else 0
            total_loc_added = user_usage.get("total_loc_added", 0) if user_usage else 0
            days_active = user_usage.get("days_active", 0) if user_usage else 0
        
        # 主要IDE（最も使用されているIDE）
        ides = user_usage.get("ides", {}) if user_usage and not usage_metrics_warning else {}
        primary_ide = max(ides.items(), key=lambda x: x[1])[0] if ides else ""

        ws.append(
            [
                name,
                login,
                joined,
                enabled,
                has_seat,
                last_disp,
                editor,
                created_disp,
                cancel_disp,
                total_interactions,
                total_generations,
                total_acceptances,
                total_loc_added,
                days_active,
                primary_ide,
            ]
        )
        reported_users += 1

        row = ws.max_row
        if inactive:
            ws.cell(row=row, column=1).font = Font(bold=True, color="FFFF0000")
            inactive_users.append(
                {
                    "氏名": name,
                    "GitHubユーザー名": login,
                    "pmed-inc参加": joined,
                    "Copilot有効/無効": enabled,
                    "シート割り当て": has_seat,
                    "最終利用日時": last_disp,
                    "最終利用エディタ": editor,
                    "シート作成日": created_disp,
                    "シート取消予定日": cancel_disp,
                }
            )

    total_users = reported_users
    inactive_count = len(inactive_users)
    active_count = max(total_users - inactive_count, 0)

    if total_users > 0:
        inactive_pct = f"{round(inactive_count / total_users * 100)}%"
        active_pct = f"{round(active_count / total_users * 100)}%"
    else:
        inactive_pct = ""
        active_pct = ""

    for _ in range(3):
        ws.append([""])

    ws.append([""])

    summary_start = ws.max_row + 1
    ws.cell(row=summary_start, column=1).value = f"【{INACTIVE_DAYS}日以上未利用ユーザー（赤字対象）一覧】"

    row = summary_start + 1
    summary_header = [
        "氏名",
        "GitHubユーザー名",
        "pmed-inc参加",
        "Copilot有効/無効",
        "シート割り当て",
        "最終利用日時",
        "最終利用エディタ",
        "シート作成日",
        "シート取消予定日",
    ]
    for col, title in enumerate(summary_header, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = title
        cell.font = Font(bold=True)

    row += 1
    for iu in inactive_users:
        name_cell = ws.cell(row=row, column=1)
        name_cell.value = iu["氏名"]
        name_cell.font = Font(bold=True, color="FFFF0000")

        ws.cell(row=row, column=2).value = iu["GitHubユーザー名"]
        ws.cell(row=row, column=3).value = iu["pmed-inc参加"]
        ws.cell(row=row, column=4).value = iu["Copilot有効/無効"]
        ws.cell(row=row, column=5).value = iu["シート割り当て"]
        ws.cell(row=row, column=6).value = iu["最終利用日時"]
        ws.cell(row=row, column=7).value = iu["最終利用エディタ"]
        ws.cell(row=row, column=8).value = iu["シート作成日"]
        ws.cell(row=row, column=9).value = iu["シート取消予定日"]

        row += 1

    row += 1
    ws.cell(row=row, column=1).value = f"{INACTIVE_DAYS}日以上未利用ユーザー数"
    ws.cell(row=row, column=2).value = inactive_count
    ws.cell(row=row, column=3).value = inactive_pct

    row += 1
    ws.cell(row=row, column=1).value = f"{INACTIVE_DAYS}日以内に利用したユーザー数"
    ws.cell(row=row, column=2).value = active_count
    ws.cell(row=row, column=3).value = active_pct

    row += 1
    ws.cell(row=row, column=1).value = "総ユーザー数"
    ws.cell(row=row, column=2).value = total_users

    logging.info("統計: 総ユーザー数=%d, 未利用=%d, 利用=%d", total_users, inactive_count, active_count)

    # ============================================================
    # 組織全体のメトリクス統計シートを追加（常に作成）
    # ============================================================
    ws_metrics = wb.create_sheet(title="組織全体の統計")
    
    # タイトル行
    ws_metrics.append([f"組織全体のCopilot使用統計 ({team_label})", f"期間: {since_date} ~ {until_date}"])
    ws_metrics.append([])
    
    # 注意事項
    ws_metrics.append(["注意: 新しい Copilot Usage Metrics API を優先使用し、不可なら legacy API にフォールバックします"])
    ws_metrics.append(["ユーザー別の詳細データは別シートに出力されます（権限またはポリシー不足時はスキップ）"])
    ws_metrics.append([])
    
    # メトリクスAPIの状態をチェック
    logging.info("組織メトリクス統計シートを作成中...")
    
    if org_metrics_data and len(org_metrics_data) > 0:
        ws_metrics.append(["メトリクスAPI呼び出し: 成功"])
        ws_metrics.append([f"取得データ: {len(org_metrics_data)} 日分"])
        ws_metrics.append([])
        
        # サマリー統計
        ws_metrics.append(["サマリー統計"])
        metrics_source = org_metrics_data[0].get("_source", "legacy_metrics")
        if metrics_source == "usage_metrics":
            total_active = sum(day.get("daily_active_users", 0) for day in org_metrics_data)
            total_interactions = sum(day.get("user_initiated_interaction_count", 0) for day in org_metrics_data)
            total_generations = sum(day.get("code_generation_activity_count", 0) for day in org_metrics_data)
            total_acceptances = sum(day.get("code_acceptance_activity_count", 0) for day in org_metrics_data)
            ws_metrics.append(["期間中の延べ日次アクティブユーザー数", total_active])
            ws_metrics.append(["期間中のユーザー起点インタラクション数", total_interactions])
            ws_metrics.append(["期間中のコード生成数", total_generations])
            ws_metrics.append(["期間中の承認数", total_acceptances])
            ws_metrics.append([])

            ws_metrics.append(["日別統計データ"])
            ws_metrics.append([
                "日付",
                "日次アクティブユーザー",
                "週次アクティブユーザー",
                "月次アクティブユーザー",
                "CLIアクティブユーザー",
                "ユーザー起点インタラクション数",
                "コード生成数",
                "承認数",
                "提案LoC",
                "追加LoC",
                "PR作成数",
                "PRマージ数",
            ])

            for day in org_metrics_data:
                ws_metrics.append([
                    day.get("date", ""),
                    day.get("daily_active_users", 0),
                    day.get("weekly_active_users", 0),
                    day.get("monthly_active_users", 0),
                    day.get("daily_active_cli_users", 0),
                    day.get("user_initiated_interaction_count", 0),
                    day.get("code_generation_activity_count", 0),
                    day.get("code_acceptance_activity_count", 0),
                    day.get("loc_suggested_to_add_sum", 0),
                    day.get("loc_added_sum", 0),
                    day.get("pull_requests_total_created", 0),
                    day.get("pull_requests_total_merged", 0),
                ])
        else:
            total_active = sum(day.get("total_active_users", 0) for day in org_metrics_data)
            total_engaged = sum(day.get("total_engaged_users", 0) for day in org_metrics_data)
            ws_metrics.append(["期間中の延べアクティブユーザー数", total_active])
            ws_metrics.append(["期間中の延べエンゲージユーザー数", total_engaged])
            ws_metrics.append([])
            
            ws_metrics.append(["日別統計データ"])
            ws_metrics.append(["日付", "アクティブユーザー数", "エンゲージユーザー数", 
                              "コード補完利用者", "コード補完提案数", "コード補完承認数",
                              "IDE Chat利用者", "IDE Chatメッセージ数", "IDE Chat挿入数"])
            
            for day in org_metrics_data:
                date = day.get("date", "")
                active_users = day.get("total_active_users", 0)
                engaged_users = day.get("total_engaged_users", 0)
                
                code_completions = day.get("copilot_ide_code_completions", {})
                code_users = code_completions.get("total_engaged_users", 0)
                code_suggestions = 0
                code_acceptances = 0
                if "editors" in code_completions:
                    for editor in code_completions["editors"]:
                        code_suggestions += editor.get("total_code_suggestions", 0)
                        code_acceptances += editor.get("total_code_acceptances", 0)
                
                ide_chat = day.get("copilot_ide_chat", {})
                chat_users = ide_chat.get("total_engaged_users", 0)
                chat_turns = 0
                chat_insertions = 0
                if "editors" in ide_chat:
                    for editor in ide_chat["editors"]:
                        chat_turns += editor.get("total_chat_turns", 0)
                        chat_insertions += editor.get("total_chat_insertion_events", 0)
                
                ws_metrics.append([date, active_users, engaged_users,
                                 code_users, code_suggestions, code_acceptances,
                                 chat_users, chat_turns, chat_insertions])
    else:
        ws_metrics.append(["メトリクスAPIへのアクセスに問題があります"])
        ws_metrics.append(["トークンの権限またはAPI設定を確認してください"])
    
    ws_metrics.append([])
    ws_metrics.append(["詳細情報"])
    ws_metrics.append(["対象期間", f"{since_date} ~ {until_date} (28日間)"])
    ws_metrics.append(["対象組織", org])
    ws_metrics.append(["対象チーム", team_label])
    ws_metrics.append(["レポート対象ユーザー数", total_users])
    
    # サマリー統計を追加
    if org_metrics_data and len(org_metrics_data) > 0:
        metrics_source = org_metrics_data[0].get("_source", "legacy_metrics")
        if metrics_source == "usage_metrics":
            total_active = sum(day.get("daily_active_users", 0) for day in org_metrics_data)
            total_engaged = sum(day.get("user_initiated_interaction_count", 0) for day in org_metrics_data)
        else:
            total_active = sum(day.get("total_active_users", 0) for day in org_metrics_data)
            total_engaged = sum(day.get("total_engaged_users", 0) for day in org_metrics_data)
        avg_active = total_active / len(org_metrics_data)
        avg_engaged = total_engaged / len(org_metrics_data)
        
        ws_metrics.append([])
        ws_metrics.append(["現在のデータ概要"])
        ws_metrics.append(["期間", f"{since_date} ~ {until_date} ({len(org_metrics_data)}日分)"])
        ws_metrics.append(["延べアクティブユーザー", f"{total_active:,}人日"])
        ws_metrics.append(["延べエンゲージユーザー", f"{total_engaged:,}人日"])
        ws_metrics.append(["1日あたり平均アクティブ", f"{avg_active:.1f}人"])
        ws_metrics.append(["1日あたり平均エンゲージ", f"{avg_engaged:.1f}人"])

    # ============================================================
    # メトリクス項目説明シートを追加
    # ============================================================
    ws_explanation = wb.create_sheet(title="メトリクス項目説明")
    
    # タイトル
    ws_explanation.append(["GitHub Copilot メトリクス項目の説明"])
    ws_explanation.append([])
    
    # 日別統計データの説明
    ws_explanation.append(["■ 日別統計データの各項目"])
    ws_explanation.append([])
    
    ws_explanation.append(["1. 日付"])
    ws_explanation.append(["", "集計対象の日（例: 2025-11-10）"])
    ws_explanation.append(["", "過去28日分の日次データ"])
    ws_explanation.append([])
    
    ws_explanation.append(["2. 日次アクティブユーザー (daily_active_users)"])
    ws_explanation.append(["", "その日にCopilotを利用したユニークユーザー数"])
    ws_explanation.append(["", "新しいUsage Metrics APIの基本的な採用指標"])
    ws_explanation.append([])
    
    ws_explanation.append(["3. ユーザー起点インタラクション数 (user_initiated_interaction_count)"])
    ws_explanation.append(["", "ユーザーがCopilotに対して能動的に行った操作回数"])
    ws_explanation.append(["", "ChatやCLIなどの利用深度を見る指標"])
    ws_explanation.append([])
    
    ws_explanation.append(["4. コード生成数 (code_generation_activity_count)"])
    ws_explanation.append(["", "Copilotが生成した提案イベント数"])
    ws_explanation.append(["", "新APIではコード補完や各機能の使用量をこの系列で見ます"])
    ws_explanation.append([])
    
    ws_explanation.append(["5. 承認数 (code_acceptance_activity_count)"])
    ws_explanation.append(["", "ユーザーが受け入れたコード生成イベント数"])
    ws_explanation.append(["", "生成数と合わせて受容度の傾向を見ます"])
    ws_explanation.append([])
    
    ws_explanation.append(["6. 提案LoC / 追加LoC"])
    ws_explanation.append(["", "提案された行数と、実際に追加された行数"])
    ws_explanation.append(["", "Copilotの出力量と実利用量を比較できます"])
    ws_explanation.append([])
    
    ws_explanation.append(["7. 週次/月次アクティブユーザー"])
    ws_explanation.append(["", "継続利用の広がりを見る採用指標です"])
    ws_explanation.append([])
    
    ws_explanation.append(["8. CLIアクティブユーザー"])
    ws_explanation.append(["", "Copilot CLI を利用したユニークユーザー数です"])
    ws_explanation.append([])
    
    ws_explanation.append(["9. PR作成数 / PRマージ数"])
    ws_explanation.append(["", "新しいUsage Metrics APIではPRライフサイクル指標も取得できます"])
    ws_explanation.append([])
    
    # 重要なポイント
    ws_explanation.append(["■ 重要なポイント"])
    ws_explanation.append([])
    ws_explanation.append(["・新APIが優先", "ユーザー別・組織別とも Copilot Usage Metrics API を優先利用します"])
    ws_explanation.append(["・Seat APIは別物", "ライセンス割当や最終利用日時は Copilot user management API が正です"])
    ws_explanation.append(["・組織集計と個人集計の用途を分離", "組織全体の傾向は統計シート、個別行動は詳細シートで確認します"])
    ws_explanation.append(["・権限とポリシー依存", "View Organization Copilot Metrics 権限と Copilot usage metrics policy が必要です"])
    ws_explanation.append([])
    
    logging.info("メトリクス項目説明シートを作成しました")

    # ============================================================
    # ユーザー別詳細メトリクスシートを追加
    # ============================================================
    ws_user_metrics = wb.create_sheet(title="ユーザー別詳細メトリクス")

    # タイトル行
    ws_user_metrics.append([f"ユーザー別Copilot使用詳細 ({team_label})", "期間: 直近28日間"])
    ws_user_metrics.append([])
    ws_user_metrics.append(["注意: このデータは Usage Metrics API から取得"])

    if usage_metrics_warning:
        ws_user_metrics.append(["取得不可", usage_metrics_warning])
        ws_user_metrics.append(["補足", "メインシートの 28 日メトリクス列は空欄で出力しています"])
        logging.info("ユーザー別詳細メトリクスシートを説明付きで作成: %s", usage_metrics_warning)
    elif usage_metrics_data and len(usage_metrics_data) > 0:
        ws_user_metrics.append(["取得成功", f"{len(usage_metrics_data)} ユーザー"])
        ws_user_metrics.append([])

        # ヘッダー
        ws_user_metrics.append([
            "ユーザー名",
            "プロンプト数",
            "コード生成数",
            "受け入れ数",
            "提案LoC",
            "追加LoC",
            "削除LoC",
            "提案削除LoC",
            "アクティブ日数",
            "Chat利用日数",
            "CLI利用日数",
            "Agent利用日数",
            "CLIプロンプト数",
            "CLIリクエスト数",
            "CLIセッション数",
            "主要機能",
            "主要IDE",
            "主要言語"
        ])

        # データ行（受け入れ数でソート）
        sorted_users = sorted(usage_metrics_data.items(),
                            key=lambda x: x[1]["total_acceptances"],
                            reverse=True)

        for username, data in sorted_users:
            ides = data.get("ides", {})
            primary_ide = max(ides.items(), key=lambda x: x[1])[0] if ides else ""

            languages = data.get("languages", {})
            primary_lang = max(languages.items(), key=lambda x: x[1])[0] if languages else ""

            features = data.get("features", {})
            primary_feature = max(features.items(), key=lambda x: x[1])[0] if features else ""

            ws_user_metrics.append([
                username,
                data.get("total_interactions", 0),
                data.get("total_generations", 0),
                data.get("total_acceptances", 0),
                data.get("total_loc_suggested", 0),
                data.get("total_loc_added", 0),
                data.get("total_loc_deleted", 0),
                data.get("total_loc_suggested_deleted", 0),
                data.get("days_active", 0),
                data.get("days_used_chat", 0),
                data.get("days_used_cli", 0),
                data.get("days_used_agent", 0),
                data.get("cli_prompt_count", 0),
                data.get("cli_request_count", 0),
                data.get("cli_session_count", 0),
                primary_feature,
                primary_ide,
                primary_lang
            ])

        logging.info("ユーザー別詳細メトリクスシートを作成: %d ユーザー", len(usage_metrics_data))
    else:
        ws_user_metrics.append(["取得結果", "ユーザー別 Usage Metrics は 0 件でした"])
        logging.info("ユーザー別詳細メトリクスシートを空データで作成しました")

    # 保存先ディレクトリ
    output_dir = "./data"
    os.makedirs(output_dir, exist_ok=True)

    # ファイル名（JSTタイムスタンプ付き）
    timestamp = now_jst.strftime("%Y%m%d%H%M%S")
    output = os.path.join(output_dir, f"copilot_report_{timestamp}.xlsx")

    # 保存
    try:
        logging.info("Excel ファイルを保存中: %s", output)
        wb.save(output)
        logging.info("保存完了")
    except PermissionError as e:
        logging.error("レポートファイル '%s' を保存できませんでした。", output)
        logging.error("Excel などでファイルが開かれていないか確認し、閉じてからもう一度実行してください。")
        logging.error("詳細: %s", e)
        sys.exit(1)
    except (OSError, IOError) as e:
        logging.error("予期しないエラーが発生しました: %s", e)
        sys.exit(1)

    logging.info("=" * 60)
    logging.info("処理完了: Excel ファイルを出力しました: %s", output)
    logging.info("デバッグログ: %s", log_file)
    logging.info("=" * 60)
    
    print(f"Excel ファイルを出力しました: {output}")
    print(f"デバッグログ: {log_file}")

if __name__ == "__main__":
    main()
