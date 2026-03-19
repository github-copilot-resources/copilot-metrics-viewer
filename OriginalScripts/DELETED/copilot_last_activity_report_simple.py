#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ============================================================
# GitHub Copilot 個別利用状況レポート生成スクリプト（Excel 出力版）
#
# 【このスクリプトがやること】
# - user.csv（氏名、GitHubユーザー名、参加区分、Copilot有効/無効）を読み込み
# - GitHub API から Copilot の最終利用日時・最終利用エディタなどを取得
# - Excel（copilot_report.xlsx）として出力
# - 最終利用から 7 日以上経過しているユーザーは「氏名を太字赤」にする
# - 末尾 3 行空けてから、
#     ・1週間以上未利用ユーザー一覧（氏名・最終利用日時・最終利用エディタ）
#     ・1週間以上未利用ユーザー数／1週間以内利用ユーザー数／総ユーザー数
#     ・1週間以上未利用ユーザー数(%)／1週間以内利用ユーザー数(%)
#   を追記する
#
# 【呼び出し方（PowerShell例）】
#
#   pip install requests openpyxl
#
#   # .env を指定して実行
#   python copilot_last_activity_report.py user.csv .env
#
#   # 第2引数（.env パス）を省略した場合はカレントディレクトリの .env を使用
#   python copilot_last_activity_report.py user.csv
#
# 【.env の例】
#
#   NUXT_GITHUB_TOKEN=ghp_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
#   NUXT_PUBLIC_GITHUB_ORG=medley-inc
#   TARGET_TEAM_SLUG=mall   # ← 任意。指定すると mall チームメンバーのみを対象にする
#
# ============================================================

import csv
import os
import sys
from typing import Dict, Any, List, Optional, Set

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timezone

GITHUB_API_BASE = "https://api.github.com"


# =========================
# 日付パース & フォーマット
# =========================
def parse_iso_datetime(dt_str: str) -> Optional[datetime]:
    """GitHub API から返る ISO8601 文字列を datetime に変換する"""
    if not dt_str:
        return None
    try:
        # "Z" を "+00:00" に変換してからパース
        return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except Exception:
        return None


def format_datetime(dt_str: str) -> str:
    """ISO8601 文字列を 'yyyy/MM/dd HH:MM' 形式に変換する"""
    dt = parse_iso_datetime(dt_str)
    if not dt:
        return ""
    return dt.strftime("%Y/%m/%d %H:%M")


# =========================
# 設定ファイル(.env)読み取り
# =========================
def load_config_from_env_file(path: str = ".env") -> (str, str, Optional[str]):
    """
    .env ファイルから
      - NUXT_GITHUB_TOKEN / GITHUB_TOKEN
      - NUXT_PUBLIC_GITHUB_ORG / GITHUB_ORG
      - TARGET_TEAM_SLUG（任意）
    を読み取る
    """
    if not os.path.exists(path):
        raise RuntimeError(f"設定ファイルが見つかりません: {path}")

    env: Dict[str, str] = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or "=" not in line or line.startswith("#"):
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip()

    token = env.get("NUXT_GITHUB_TOKEN") or env.get("GITHUB_TOKEN")
    org = env.get("NUXT_PUBLIC_GITHUB_ORG") or env.get("GITHUB_ORG")
    team_slug = env.get("TARGET_TEAM_SLUG")  # 任意

    if not token:
        raise RuntimeError("設定ファイルに NUXT_GITHUB_TOKEN または GITHUB_TOKEN がありません。")
    if not org:
        raise RuntimeError("設定ファイルに NUXT_PUBLIC_GITHUB_ORG または GITHUB_ORG がありません。")

    return token, org, team_slug


def github_headers(token: str) -> Dict[str, str]:
    """GitHub API 呼び出し用ヘッダ"""
    return {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
    }


# =========================
# GitHub API：シート一覧取得
# =========================
def fetch_all_seats(org: str, token: str) -> List[Dict[str, Any]]:
    """
    /orgs/{org}/copilot/billing/seats を叩いて
    Copilot シート情報を全件取得（ページング対応）
    """
    seats: List[Dict[str, Any]] = []
    url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/billing/seats"
    headers = github_headers(token)
    params: Optional[Dict[str, Any]] = {"per_page": 100}

    while url:
        resp = requests.get(url, headers=headers, params=params)
        resp.raise_for_status()
        data = resp.json()

        seats.extend(data.get("seats", []))

        # Link ヘッダから next ページの URL を探す
        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            for part in link.split(","):
                part = part.strip()
                if 'rel="next"' in part:
                    start = part.find("<") + 1
                    end = part.find(">")
                    next_url = part[start:end]
                    break

        url = next_url
        params = None  # 2ページ目以降は URL にクエリが含まれている想定

    return seats


# =========================
# GitHub API：チームメンバー取得
# =========================
def fetch_team_members(org: str, token: str, team_slug: str) -> Set[str]:
    """
    /orgs/{org}/teams/{team_slug}/members を叩いて
    チームに所属するメンバーの login 一覧を取得する（ページング対応）
    """
    members: Set[str] = set()
    if not team_slug:
        return members

    url = f"{GITHUB_API_BASE}/orgs/{org}/teams/{team_slug}/members"
    headers = github_headers(token)
    params: Optional[Dict[str, Any]] = {"per_page": 100}

    while url:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code == 404:
            # チームが存在しない場合など
            break
        resp.raise_for_status()
        data = resp.json()

        for m in data:
            login = m.get("login")
            if login:
                members.add(login)

        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            for part in link.split(","):
                part = part.strip()
                if 'rel="next"' in part:
                    start = part.find("<") + 1
                    end = part.find(">")
                    next_url = part[start:end]
                    break

        url = next_url
        params = None

    return members


# =========================
# GitHub API：個別ユーザー詳細
# =========================
def fetch_user_seat_details(org: str, token: str, username: str) -> Optional[Dict[str, Any]]:
    """
    /orgs/{org}/members/{username}/copilot を叩いて
    個別ユーザーの Copilot 利用詳細を取得する
    """
    url = f"{GITHUB_API_BASE}/orgs/{org}/members/{username}/copilot"
    headers = github_headers(token)
    resp = requests.get(url, headers=headers)

    if resp.status_code in (404, 422):
        # シート未割り当てなど
        return None

    resp.raise_for_status()
    return resp.json()


# =========================
# CSV 読み込み
# =========================
def read_user_csv(path: str) -> List[Dict[str, str]]:
    """ユーザー一覧 CSV を読み込む"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# =========================
# メイン処理
# =========================
def main() -> None:
    if len(sys.argv) < 2:
        print("使い方: python copilot_last_activity_report.py user.csv [.env]", file=sys.stderr)
        sys.exit(1)

    user_csv_path = sys.argv[1]
    config_path = sys.argv[2] if len(sys.argv) >= 3 else ".env"

    # 設定ロード
    token, org, team_slug = load_config_from_env_file(config_path)

    # チームメンバー取得（TARGET_TEAM_SLUG が設定されている場合のみ）
    team_members: Optional[Set[str]] = None
    if team_slug:
        team_members = fetch_team_members(org, token, team_slug)
        print(f"TARGET_TEAM_SLUG={team_slug} が指定されました。チームメンバー {len(team_members)} 名を対象とします。")

    # Copilot シート情報取得
    seats = fetch_all_seats(org, token)
    seat_by_login: Dict[str, Dict[str, Any]] = {
        seat.get("assignee", {}).get("login"): seat
        for seat in seats
        if seat.get("assignee", {}).get("login")
    }

    # ユーザー CSV 読み込み
    users = read_user_csv(user_csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Copilot Report"

    header = [
        "氏名",
        "GitHubユーザー名",
        "pmed-inc参加",
        "Copilot有効/無効",
        "シート割り当て",
        "最終利用日時",
        "最終利用エディタ",
        "シート作成日",
        "シート取消予定日",
    ]
    ws.append(header)

    now_utc = datetime.now(timezone.utc)

    # 7日以上未利用 or シートなしのユーザー一覧（サマリー用）
    inactive_users: List[Dict[str, str]] = []

    # ===== メイン一覧生成 =====
    for u in users:
        name = u.get("氏名", "")
        login = u.get("GitHubユーザー名", "")
        joined = u.get("pmed-incへの参加", "")
        enabled = u.get("GitHub Copilotの有効/無効", "")

        # チーム指定がある場合は、team_members に含まれないユーザーはスキップ
        if team_members is not None and login not in team_members:
            continue

        seat = seat_by_login.get(login)
        has_seat = "あり" if seat else "なし"

        last_disp = ""
        created_disp = ""
        cancel_disp = ""
        editor = ""
        inactive_7 = False

        if seat:
            raw_last = seat.get("last_activity_at") or ""
            raw_created = seat.get("created_at") or ""
            raw_cancel = seat.get("pending_cancellation_date") or ""

            last_disp = format_datetime(raw_last)
            created_disp = format_datetime(raw_created)
            cancel_disp = format_datetime(raw_cancel)

            details = fetch_user_seat_details(org, token, login)
            editor = details.get("last_activity_editor") if details else ""

            last_dt = parse_iso_datetime(raw_last)
            if last_dt is None:
                inactive_7 = True
            else:
                delta = now_utc - last_dt.astimezone(timezone.utc)
                if delta.days >= 7:
                    inactive_7 = True
        else:
            # シート自体が無い場合も未利用扱いとする
            inactive_7 = True

        ws.append([
            name,
            login,
            joined,
            enabled,
            has_seat,
            last_disp,
            editor,
            created_disp,
            cancel_disp,
        ])

        # 行ハイライトと inactive_users への追加
        row = ws.max_row
        if inactive_7:
            ws.cell(row=row, column=1).font = Font(bold=True, color="FFFF0000")
            inactive_users.append(
                {
                    "name": name,
                    "last": last_disp,
                    "editor": editor,
                }
            )

    # ===== サマリー生成 =====

    # 対象ユーザー数（team_slug ありの場合はそのチームに属する人数）
    total_users = len({u["name"] for u in inactive_users})  # inactive だけではなく、本来は全対象ユーザー数
    # 上の total_users は inactive_users から計算すると「未利用者数」になってしまうので、
    # team_members or users ベースで再計算する
    if team_members is not None:
        # team_members に含まれる user.csv 上のユーザー数（実際にレポート出力した人数）
        reported_logins = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            login_val = row[0].value
            if login_val:
                reported_logins.add(login_val)
        total_users = len(reported_logins)
    else:
        # チーム指定なしの場合は user.csv の件数
        total_users = len(users)

    inactive_count = len(inactive_users)
    active_count = max(total_users - inactive_count, 0)

    if total_users > 0:
        inactive_pct = f"{round(inactive_count / total_users * 100)}%"
        active_pct = f"{round(active_count / total_users * 100)}%"
    else:
        inactive_pct = ""
        active_pct = ""

    # メイン表の末尾に 3 行空ける
    for _ in range(3):
        ws.append([""])

    # サマリータイトルの前に 1 行空行
    ws.append([""])

    # サマリー開始行（タイトル行）
    summary_start = ws.max_row + 1
    ws.cell(row=summary_start, column=1).value = "【1週間以上未利用ユーザー（赤字対象）一覧】"

    # 一覧のヘッダー行
    row = summary_start + 1
    ws.cell(row=row, column=1).value = "氏名"
    ws.cell(row=row, column=2).value = "最終利用日時"
    ws.cell(row=row, column=3).value = "最終利用エディタ"
    for col in range(1, 4):
        ws.cell(row=row, column=col).font = Font(bold=True)

    # 赤字ユーザー一覧
    row += 1
    for iu in inactive_users:
        name_cell = ws.cell(row=row, column=1)
        name_cell.value = iu["name"]
        name_cell.font = Font(bold=True, color="FFFF0000")

        ws.cell(row=row, column=2).value = iu["last"]
        ws.cell(row=row, column=3).value = iu["editor"]

        row += 1

    # サマリー数値
    row += 1
    ws.cell(row=row, column=1).value = "1週間以上未利用ユーザー数"
    ws.cell(row=row, column=2).value = inactive_count
    ws.cell(row=row, column=3).value = inactive_pct

    row += 1
    ws.cell(row=row, column=1).value = "1週間以内に利用したユーザー数"
    ws.cell(row=row, column=2).value = active_count
    ws.cell(row=row, column=3).value = active_pct

    row += 1
    ws.cell(row=row, column=1).value = "総ユーザー数"
    ws.cell(row=row, column=2).value = total_users

    # 保存（Excel で開かれている場合のエラーをハンドリング）
    output = "copilot_report.xlsx"
    try:
        wb.save(output)
    except PermissionError as e:
        print(
            f"レポートファイル '{output}' を保存できませんでした。",
            file=sys.stderr,
        )
        print(
            "Excel などでファイルが開かれていないか確認し、閉じてからもう一度実行してください。",
            file=sys.stderr,
        )
        print(f"詳細: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Excel ファイルを出力しました: {output}")


if __name__ == "__main__":
    main()
