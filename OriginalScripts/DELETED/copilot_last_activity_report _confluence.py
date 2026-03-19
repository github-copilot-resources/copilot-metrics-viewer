#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ============================================================
# GitHub Copilot 個別利用状況レポート生成スクリプト（Excel 出力版）
#
# 【このスクリプトがやること】
# - user.csv（氏名、GitHubユーザー名、参加区分、Copilot有効/無効）を読み込み
# - GitHub API から Copilot の最終利用日時・最終利用エディタなどを取得
# - Excel（copilot_report.xlsx）として出力
# - 最終利用から 7 日以上経過しているユーザーは「氏名を太字赤」にする
# - 末尾 3 行空けてから、
#     ・1週間以上未利用ユーザー一覧（既存と同じ9項目を出力）
#     ・1週間以上未利用ユーザー数／1週間以内利用ユーザー数／総ユーザー数
#     ・それぞれの割合(%) を出力する
#
# 【チーム絞り込み】
# - .env に TARGET_TEAM_SLUG=mall を指定すると、
#   GitHub の "mall" チームメンバーに含まれるユーザーのみを対象にレポートする
# - TARGET_TEAM_SLUG が未指定の場合は、user.csv に記載された全ユーザーを対象とする
#
# 【呼び出し方（PowerShell例）】
#
#   pip install requests openpyxl
#
#   python copilot_last_activity_report.py user.csv .env
#
# 【.env の例】
#
#   NUXT_GITHUB_TOKEN=ghp_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
#   NUXT_PUBLIC_GITHUB_ORG=medley-inc
#   TARGET_TEAM_SLUG=mall   # 任意。指定すると mall チームメンバーのみを対象にする
#
# ============================================================

import csv
import os
import sys
from typing import Dict, Any, List, Optional, Set

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timezone

GITHUB_API_BASE = "https://api.github.com"


# =========================
# 日付パース & フォーマット
# =========================
def parse_iso_datetime(dt_str: str) -> Optional[datetime]:
    """GitHub API の ISO8601 文字列を datetime に変換"""
    if not dt_str:
        return None
    try:
        # 例: "2024-01-01T12:34:56Z"
        return datetime.strptime(dt_str, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def format_datetime(dt_str: str) -> str:
    """ISO8601 文字列を 'YYYY/MM/DD HH:MM' 形式に整形（パースできない場合は空文字）"""
    dt = parse_iso_datetime(dt_str)
    if not dt:
        return ""
    return dt.strftime("%Y/%m/%d %H:%M")


# =========================
# 設定ファイル(.env)読み取り
# =========================
def load_config_from_env_file(path: str = ".env") -> (str, str, Optional[str]):
    """
    .env ファイルから
      - NUXT_GITHUB_TOKEN / GITHUB_TOKEN
      - NUXT_PUBLIC_GITHUB_ORG / GITHUB_ORG
      - TARGET_TEAM_SLUG（任意）
    を読み込む。
    """
    token = None
    org = None
    team_slug: Optional[str] = None

    # まずは .env ファイルを読む
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if k == "NUXT_GITHUB_TOKEN" or k == "GITHUB_TOKEN":
                    token = v
                elif k == "NUXT_PUBLIC_GITHUB_ORG" or k == "GITHUB_ORG":
                    org = v
                elif k == "TARGET_TEAM_SLUG":
                    team_slug = v

    # 環境変数を上書き優先
    token = os.environ.get("NUXT_GITHUB_TOKEN") or os.environ.get("GITHUB_TOKEN") or token
    org = os.environ.get("NUXT_PUBLIC_GITHUB_ORG") or os.environ.get("GITHUB_ORG") or org
    team_slug = os.environ.get("TARGET_TEAM_SLUG") or team_slug

    if not token:
        print(
            "GitHub Token が見つかりません。.env または環境変数 NUXT_GITHUB_TOKEN / GITHUB_TOKEN を確認してください。",
            file=sys.stderr,
        )
        sys.exit(1)
    if not org:
        print(
            "GitHub Organization が見つかりません。.env または環境変数 NUXT_PUBLIC_GITHUB_ORG / GITHUB_ORG を確認してください。",
            file=sys.stderr,
        )
        sys.exit(1)

    return token, org, team_slug


# =========================
# GitHub API 呼び出し共通
# =========================
def github_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


# =========================
# Copilot シート情報取得
# =========================
def fetch_all_seats(org: str, token: str) -> List[Dict[str, Any]]:
    """
    /orgs/{org}/copilot/billing/seats を叩いて
    シート情報一覧（複数ページ）を全て取得する
    """
    seats: List[Dict[str, Any]] = []

    url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/billing/seats"
    params: Optional[Dict[str, Any]] = {"per_page": 100}

    while url:
        resp = requests.get(url, headers=github_headers(token), params=params)
        resp.raise_for_status()
        data = resp.json()

        seats.extend(data.get("seats", []))

        # Link ヘッダから next ページの URL を探す
        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            for part in link.split(","):
                part = part.strip()
                if 'rel="next"' in part:
                    # <https://api.github.com/...>; rel="next"
                    start = part.find("<") + 1
                    end = part.find(">")
                    next_url = part[start:end]
                    break

        url = next_url
        params = None  # 2ページ目以降はクエリ不要

    return seats


# =========================
# チームメンバー一覧取得
# =========================
def fetch_team_members(org: str, token: str, team_slug: str) -> Set[str]:
    """
    /orgs/{org}/teams/{team_slug}/members を叩いて
    チームメンバー（login）の集合を返す
    """
    members: Set[str] = set()

    url = f"{GITHUB_API_BASE}/orgs/{org}/teams/{team_slug}/members"
    params: Optional[Dict[str, Any]] = {"per_page": 100}

    while url:
        resp = requests.get(url, headers=github_headers(token), params=params)
        if resp.status_code == 404:
            print(
                f"指定されたチームが見つかりません: org={org}, team_slug={team_slug}",
                file=sys.stderr,
            )
            break

        resp.raise_for_status()
        data = resp.json()

        for m in data:
            login = m.get("login")
            if login:
                members.add(login)

        link = resp.headers.get("Link", "")
        next_url = None
        if link:
            for part in link.split(","):
                part = part.strip()
                if 'rel="next"' in part:
                    start = part.find("<") + 1
                    end = part.find(">")
                    next_url = part[start:end]
                    break
        url = next_url
        params = None

    return members


# =========================
# GitHub API：個別ユーザー詳細
# =========================
def fetch_user_seat_details(org: str, token: str, username: str) -> Optional[Dict[str, Any]]:
    """
    /orgs/{org}/members/{username}/copilot を叩いて
    個別ユーザーの Copilot 利用詳細を取得する
    """
    url = f"{GITHUB_API_BASE}/orgs/{org}/members/{username}/copilot"
    headers = github_headers(token)
    resp = requests.get(url, headers=headers)

    if resp.status_code in (404, 422):
        # シート未割り当てなど
        return None

    resp.raise_for_status()
    return resp.json()


# =========================
# CSV 読み込み
# =========================
def read_user_csv(path: str) -> List[Dict[str, str]]:
    """ユーザー一覧 CSV を読み込む"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# =========================
# Markdownレポート生成
# =========================
def generate_markdown_report(
    summary: Dict[str, Any],
    template_path: str = "copilot_confluence_template.md",
    output_path: str = "copilot_confluence_report.md",
) -> None:
    """
    Confluence に貼り付ける Markdown レポートを出力する。
    テンプレート中の {{KEY}} を summary[KEY] で置換するシンプル実装。
    """
    if not os.path.exists(template_path):
        print(f"Markdownテンプレートが見つかりません: {template_path}", file=sys.stderr)
        return

    with open(template_path, "r", encoding="utf-8") as f:
        text = f.read()

    for key, value in summary.items():
        placeholder = "{{" + key + "}}"
        text = text.replace(placeholder, str(value))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text)

    print(f"Markdownレポートを出力しました: {output_path}")


# =========================
# メイン処理
# =========================
def main() -> None:
    if len(sys.argv) < 2:
        print("使い方: python copilot_last_activity_report.py user.csv [.env]", file=sys.stderr)
        sys.exit(1)

    user_csv_path = sys.argv[1]
    config_path = sys.argv[2] if len(sys.argv) >= 3 else ".env"

    # 設定ロード
    token, org, team_slug = load_config_from_env_file(config_path)

    # チームメンバー取得（TARGET_TEAM_SLUG が設定されている場合のみ）
    team_members: Optional[Set[str]] = None
    if team_slug:
        team_members = fetch_team_members(org, token, team_slug)
        print(f"TARGET_TEAM_SLUG={team_slug} が指定されました。チームメンバー {len(team_members)} 名を対象とします。")

    # Copilot シート情報取得
    seats = fetch_all_seats(org, token)
    seat_by_login: Dict[str, Dict[str, Any]] = {
        seat.get("assignee", {}).get("login"): seat
        for seat in seats
        if seat.get("assignee", {}).get("login")
    }

    # ユーザー一覧の読み込み
    users = read_user_csv(user_csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "copilot_report"

    # ヘッダー
    header = [
        "氏名",
        "GitHubユーザー名",
        "pmed-inc参加",
        "Copilot有効/無効",
        "シート割り当て",
        "最終利用日時",
        "最終利用エディタ",
        "シート作成日",
        "シート取消予定日",
    ]
    ws.append(header)

    now_utc = datetime.now(timezone.utc)

    # 7日以上未利用 or シートなしのユーザー一覧（サマリー用）
    inactive_users: List[Dict[str, str]] = []

    reported_users = 0

    # ユーザーごとのレポート行出力
    for user in users:
        name = user.get("氏名", "")
        login = user.get("GitHubユーザー名", "")
        joined = user.get("pmed-inc参加", "")
        enabled = user.get("Copilot有効/無効", "")

        if not login:
            continue

        # チーム絞り込み：team_members が指定されている場合は、その集合に含まれる login のみ対象にする
        if team_members is not None and login not in team_members:
            continue

        seat = seat_by_login.get(login)
        has_seat = "あり" if seat else "なし"

        inactive_7 = False

        # seat がある場合のみ、最終利用日時などを参照する
        last_disp = ""
        created_disp = ""
        cancel_disp = ""
        editor = ""

        if seat:
            raw_last = seat.get("last_activity_at") or ""
            raw_created = seat.get("created_at") or ""
            raw_cancel = seat.get("pending_cancellation_date") or ""

            last_disp = format_datetime(raw_last)
            created_disp = format_datetime(raw_created)
            cancel_disp = format_datetime(raw_cancel)

            details = fetch_user_seat_details(org, token, login)
            editor = details.get("last_activity_editor") if details else ""

            last_dt = parse_iso_datetime(raw_last)
            if last_dt is None:
                inactive_7 = True
            else:
                delta = now_utc - last_dt.astimezone(timezone.utc)
                if delta.days >= 7:
                    inactive_7 = True
        else:
            # シート自体が無い場合も未利用扱いとする
            inactive_7 = True

        ws.append(
            [
                name,
                login,
                joined,
                enabled,
                has_seat,
                last_disp,
                editor,
                created_disp,
                cancel_disp,
            ]
        )
        reported_users += 1

        # 行ハイライトと inactive_users への追加（9項目フルで保持）
        row = ws.max_row
        if inactive_7:
            ws.cell(row=row, column=1).font = Font(bold=True, color="FFFF0000")
            inactive_users.append(
                {
                    "氏名": name,
                    "GitHubユーザー名": login,
                    "pmed-inc参加": joined,
                    "Copilot有効/無効": enabled,
                    "シート割り当て": has_seat,
                    "最終利用日時": last_disp,
                    "最終利用エディタ": editor,
                    "シート作成日": created_disp,
                    "シート取消予定日": cancel_disp,
                }
            )

    # ===== サマリー生成 =====

    total_users = reported_users
    inactive_count = len(inactive_users)
    active_count = max(total_users - inactive_count, 0)

    if total_users > 0:
        inactive_pct = f"{round(inactive_count / total_users * 100)}%"
        active_pct = f"{round(active_count / total_users * 100)}%"
    else:
        inactive_pct = ""
        active_pct = ""

    # メイン表の末尾に 3 行空ける
    for _ in range(3):
        ws.append([""])

    # サマリータイトルの前に 1 行空行
    ws.append([""])

    # サマリー開始行（タイトル行）
    summary_start = ws.max_row + 1
    ws.cell(row=summary_start, column=1).value = "【1週間以上未利用ユーザー（赤字対象）一覧】"

    # 一覧のヘッダー行（9項目）
    row = summary_start + 1
    summary_header = [
        "氏名",
        "GitHubユーザー名",
        "pmed-inc参加",
        "Copilot有効/無効",
        "シート割り当て",
        "最終利用日時",
        "最終利用エディタ",
        "シート作成日",
        "シート取消予定日",
    ]
    for col, title in enumerate(summary_header, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = title
        cell.font = Font(bold=True)

    # 赤字ユーザー一覧（9項目フル出力）
    row += 1
    for iu in inactive_users:
        name_cell = ws.cell(row=row, column=1)
        name_cell.value = iu["氏名"]
        name_cell.font = Font(bold=True, color="FFFF0000")

        ws.cell(row=row, column=2).value = iu["GitHubユーザー名"]
        ws.cell(row=row, column=3).value = iu["pmed-inc参加"]
        ws.cell(row=row, column=4).value = iu["Copilot有効/無効"]
        ws.cell(row=row, column=5).value = iu["シート割り当て"]
        ws.cell(row=row, column=6).value = iu["最終利用日時"]
        ws.cell(row=row, column=7).value = iu["最終利用エディタ"]
        ws.cell(row=row, column=8).value = iu["シート作成日"]
        ws.cell(row=row, column=9).value = iu["シート取消予定日"]

        row += 1

    # サマリー数値
    row += 1
    ws.cell(row=row, column=1).value = "1週間以上未利用ユーザー数"
    ws.cell(row=row, column=2).value = inactive_count
    ws.cell(row=row, column=3).value = inactive_pct

    row += 1
    ws.cell(row=row, column=1).value = "1週間以内に利用したユーザー数"
    ws.cell(row=row, column=2).value = active_count
    ws.cell(row=row, column=3).value = active_pct

    row += 1
    ws.cell(row=row, column=1).value = "総ユーザー数"
    ws.cell(row=row, column=2).value = total_users

    # 保存（Excel で開かれている場合のエラーをハンドリング）
    output = "copilot_report.xlsx"
    try:
        wb.save(output)
    except PermissionError as e:
        print(
            f"レポートファイル '{output}' を保存できませんでした。",
            file=sys.stderr,
        )
        print(
            "Excel などでファイルが開かれていないか確認し、閉じてからもう一度実行してください。",
            file=sys.stderr,
        )
        print(f"詳細: {e}", file=sys.stderr)
        sys.exit(1)

    print(f"Excel ファイルを出力しました: {output}")

    # Markdown レポート生成（Confluence貼り付け用）
    team_label = team_slug if team_slug else "全体"
    summary = {
        "DATE": datetime.now().strftime("%Y-%m-%d"),
        "TARGET_TEAM": team_label,
        "TOTAL_USERS": total_users,
        "ACTIVE_COUNT": active_count,
        "INACTIVE_COUNT": inactive_count,
        "ACTIVE_PCT": active_pct,
        "INACTIVE_PCT": inactive_pct,
    }
    generate_markdown_report(summary)


if __name__ == "__main__":
    main()
