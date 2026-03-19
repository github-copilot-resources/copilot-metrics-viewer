#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ============================================================
# GitHub Copilot 利用状況サマリーを Excel に1行追記するスクリプト
#
# 【このスクリプトがやること】
# - user.csv（氏名、GitHubユーザー名、参加区分、Copilot有効/無効）を読み込み
# - GitHub API から Copilot 座席情報を取得
# - 「最終利用から7日以上経過 or シート無し」を「1週間以上未利用」とみなす
# - 以下の5列を Excel（copilot_summary_log.xlsxなど）に1行追記する：
#     データ取得日時
#     1週間以上未利用ユーザー数
#     1週間以上未利用ユーザー数(%)
#     1週間以内に利用したユーザー数
#     1週間以内に利用したユーザー数(%)
#
# 【チーム絞り込み（mall向け）】
# - .env に TARGET_TEAM_SLUG=mall を指定すると、
#   GitHub の "mall" チームメンバーに含まれるユーザーのみを対象に集計する
# - TARGET_TEAM_SLUG が未指定の場合は、user.csv に記載された全ユーザーを対象とする
#
# 【呼び出し方（PowerShell例）】
#
#   pip install requests openpyxl
#
#   python copilot_summary_append.py user.csv .env copilot_summary_log.xlsx
#
# 【.env の例】
#
#   NUXT_GITHUB_TOKEN=ghp_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
#   NUXT_PUBLIC_GITHUB_ORG=medley-inc
#   TARGET_TEAM_SLUG=mall   # 任意。指定すると mall チームメンバーのみを対象にする
#
# ============================================================
# 【キャッシュとETagについて】
# ------------------------------------------------------------
# このスクリプトは GitHub API の呼び出し回数を削減するため、
# ETagベースのキャッシュ機構を実装しています。
#
# ■ キャッシュの動作
#
# ETagベースのキャッシュ（Copilotシート情報、チームメンバー）
#    - GitHub API の ETag ヘッダーを利用した変更検知
#    - If-None-Match ヘッダーで前回の ETag を送信
#    - データ未変更時: 304 Not Modified レスポンス（軽量、キャッシュ使用）
#    - データ変更時: 200 OK レスポンス（新データ取得、キャッシュ更新）
#    - メリット: データが変更された瞬間に検知、時間制限なし
#
# ■ キャッシュファイルの保存場所
#    ./cache ディレクトリ（自動作成）
#    - JSON形式で保存
#    - ファイル名: MD5ハッシュ（org名、チーム名などから生成）
#
# ■ API呼び出し削減効果（チーム指定の場合）
#    初回実行: 2回のAPI呼び出し（シート情報 + チームメンバー）
#    2回目以降（データ未変更時）: 2回（両方とも304レスポンス、軽量）
#    データ転送量の大幅削減により、レスポンス速度が向上
#
# ■ キャッシュの無効化
#    ./cache ディレクトリを削除すると、全てのキャッシュがクリアされます
#
# ■ 注意事項
#    このスクリプトはユーザー詳細情報を取得しないため、
#    API呼び出し数は少なく、キャッシュの効果は主にデータ転送量削減です。
#
# ============================================================
# 【ChangeLog】
# 2025-12-05: コード品質改善
#   - マジックナンバーの定数化 (INACTIVE_DAYS=7, JST_OFFSET=9)
#   - ログ機能追加 (logging モジュール、処理進捗の可視化)
#   - ページネーション処理の共通化 (paginate_github_api 関数)
#   - エラーハンドリング強化 (timeout=30秒、具体的な例外型、lazy formatting)
#   - 型ヒント改善 (Tuple型の明示化)
#   - ETagベースのキャッシュ機構実装
#     * GitHub API の ETag を利用した変更検知
#     * 304 Not Modified レスポンスによる API 呼び出し削減
#     * ./cache ディレクトリにキャッシュデータを保存
#     * データ変更時のみ新規取得、未変更時はキャッシュ使用
# ============================================================

import csv
import hashlib
import json
import logging
import os
import sys
import time
from typing import Dict, Any, List, Optional, Set, Tuple

import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone, timedelta

GITHUB_API_BASE = "https://api.github.com"
INACTIVE_DAYS = 7  # 未利用と判定する日数
JST_OFFSET = 9  # JSTのオフセット（UTC+9）

# キャッシュ設定
CACHE_DIR = "./cache"
CACHE_EXPIRY_MINUTES = 10  # キャッシュ有効期限（分）

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


# =========================
# キャッシュユーティリティ
# =========================
def get_cache_key(prefix: str, *args: Any) -> str:
    """キャッシュキーを生成"""
    key_str = f"{prefix}_{'_'.join(str(arg) for arg in args)}"
    return hashlib.md5(key_str.encode()).hexdigest()


def load_cache(cache_key: str) -> tuple[Optional[Any], Optional[str]]:
    """キャッシュからデータとETagを読み込む"""
    if not os.path.exists(CACHE_DIR):
        return None, None
    
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    if not os.path.exists(cache_file):
        return None, None
    
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
        
        cached_at = cache_data.get('cached_at', 0)
        etag = cache_data.get('etag')
        data = cache_data.get('data')
        
        now = time.time()
        age_minutes = (now - cached_at) / 60
        
        logging.info("キャッシュが存在: %s (取得から%.1f分経過, ETag: %s)", 
                     cache_key, age_minutes, etag or "N/A")
        
        return data, etag
    
    except (json.JSONDecodeError, IOError) as e:
        logging.warning("キャッシュ読み込みエラー: %s - %s", cache_key, e)
        return None, None


def save_cache(cache_key: str, data: Any, etag: Optional[str] = None) -> None:
    """データとETagをキャッシュに保存"""
    os.makedirs(CACHE_DIR, exist_ok=True)
    
    cache_file = os.path.join(CACHE_DIR, f"{cache_key}.json")
    cache_data = {
        'cached_at': time.time(),
        'etag': etag,
        'data': data
    }
    
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        logging.debug("キャッシュに保存: %s (ETag: %s)", cache_key, etag or "N/A")
    except IOError as e:
        logging.warning("キャッシュ保存エラー: %s - %s", cache_key, e)


# =========================
# 日付パース
# =========================
def parse_iso_datetime(dt_str: str) -> Optional[datetime]:
    """GitHub API から返る ISO8601 文字列を datetime に変換する"""
    if not dt_str:
        return None
    try:
        # "Z" を "+00:00" に変換してからパース
        return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return None


# =========================
# 設定ファイル(.env)読み取り
# =========================
def load_config_from_env_file(path: str = ".env") -> Tuple[str, str, Optional[str]]:
    """
    .env ファイルから
      - NUXT_GITHUB_TOKEN / GITHUB_TOKEN
      - NUXT_PUBLIC_GITHUB_ORG / GITHUB_ORG
      - TARGET_TEAM_SLUG（任意）
    を読み取る
    """
    if not os.path.exists(path):
        raise RuntimeError(f"設定ファイルが見つかりません: {path}")

    env: Dict[str, str] = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or "=" not in line or line.startswith("#"):
                continue
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip()

    token = env.get("NUXT_GITHUB_TOKEN") or env.get("GITHUB_TOKEN")
    org = env.get("NUXT_PUBLIC_GITHUB_ORG") or env.get("GITHUB_ORG")
    team_slug = env.get("TARGET_TEAM_SLUG")  # 任意

    if not token:
        raise RuntimeError("設定ファイルに NUXT_GITHUB_TOKEN または GITHUB_TOKEN がありません。")
    if not org:
        raise RuntimeError("設定ファイルに NUXT_PUBLIC_GITHUB_ORG または GITHUB_ORG がありません。")

    return token, org, team_slug


def github_headers(token: str) -> Dict[str, str]:
    """GitHub API 呼び出し用ヘッダ"""
    return {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
    }


# =========================
# GitHub API ページネーション共通処理
# =========================
def paginate_github_api(url: str, token: str, data_key: Optional[str] = None, etag: Optional[str] = None) -> tuple[List[Dict[str, Any]], Optional[str]]:
    """
    GitHub API のページネーションを処理して全データを取得 (ETag対応)
    
    Args:
        url: API エンドポイント URL
        token: GitHub トークン
        data_key: レスポンスからデータを取得するキー（Noneの場合はレスポンス自体がリスト）
        etag: 前回取得時のETag
    
    Returns:
        (データリスト, 新しいETag)
    """
    results: List[Dict[str, Any]] = []
    params: Optional[Dict[str, Any]] = {"per_page": 100}
    new_etag: Optional[str] = None
    
    while url:
        try:
            headers = github_headers(token)
            # ETagがあればIf-None-Matchヘッダを追加
            if etag and not results:  # 最初のリクエストのみ
                headers['If-None-Match'] = etag
            
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            
            # 304 Not Modified - データ未変更
            if resp.status_code == 304:
                logging.info("データ未変更 (304 Not Modified) - キャッシュを使用")
                return [], etag  # 空リストと旧ETagを返す（呼び出し側でキャッシュ使用）
            
            resp.raise_for_status()
            
            # 新しいETagを取得（最初のレスポンスのみ）
            if not new_etag:
                new_etag = resp.headers.get('ETag')
            
            data = resp.json()
            
            if data_key:
                results.extend(data.get(data_key, []))
            else:
                if isinstance(data, list):
                    results.extend(data)
                else:
                    results.append(data)
            
            # 次ページのURLを取得
            link = resp.headers.get("Link", "")
            next_url = None
            if link:
                for part in link.split(","):
                    part = part.strip()
                    if 'rel="next"' in part:
                        start = part.find("<") + 1
                        end = part.find(">")
                        next_url = part[start:end]
                        break
            
            url = next_url
            params = None
            
        except requests.exceptions.RequestException as e:
            logging.error("API リクエストエラー: %s - %s", url, e)
            break
    
    return results, new_etag


# =========================
# GitHub API：シート一覧取得
# =========================
def fetch_all_seats(org: str, token: str) -> List[Dict[str, Any]]:
    """
    /orgs/{org}/copilot/billing/seats を叩いて
    Copilot シート情報を全件取得（ページング対応、ETagキャッシュ対応）
    """
    cache_key = get_cache_key("seats", org)
    
    # キャッシュチェック
    cached_data, cached_etag = load_cache(cache_key)
    
    # API呼び出し
    url = f"{GITHUB_API_BASE}/orgs/{org}/copilot/billing/seats"
    logging.info("Copilot シート情報を取得中: %s (ETag付きリクエスト)", org)
    
    seats, new_etag = paginate_github_api(url, token, data_key="seats", etag=cached_etag)
    
    # 304の場合はキャッシュを使用
    if not seats and cached_data is not None:
        logging.info("キャッシュデータを使用: %d件", len(cached_data))
        return cached_data
    
    # 新しいデータをキャッシュに保存
    if seats:
        logging.info("新しいデータを取得: %d件", len(seats))
        save_cache(cache_key, seats, new_etag)
        return seats
    
    # キャッシュも新データもない場合
    return []


# =========================
# GitHub API：チームメンバー取得
# =========================
def fetch_team_members(org: str, token: str, team_slug: str) -> Set[str]:
    """
    /orgs/{org}/teams/{team_slug}/members を叩いて
    チームに所属するメンバーの login 一覧を取得する（ページング対応、ETagキャッシュ対応）
    """
    if not team_slug:
        return set()
    
    cache_key = get_cache_key("team_members", org, team_slug)
    
    # キャッシュチェック
    cached_data, cached_etag = load_cache(cache_key)
    
    url = f"{GITHUB_API_BASE}/orgs/{org}/teams/{team_slug}/members"
    logging.info("チームメンバーを取得中: %s (ETag付きリクエスト)", team_slug)
    
    try:
        # 404チェックのため最初のリクエストを個別に実行
        headers = github_headers(token)
        if cached_etag:
            headers['If-None-Match'] = cached_etag
        
        resp = requests.get(url, headers=headers, params={"per_page": 100}, timeout=30)
        
        if resp.status_code == 404:
            logging.error("指定されたチームが見つかりません: org=%s, team_slug=%s", org, team_slug)
            return set()
        
        # 304 Not Modified - キャッシュ使用
        if resp.status_code == 304:
            logging.info("データ未変更 (304) - キャッシュを使用: %d名", len(cached_data) if cached_data else 0)
            return set(cached_data) if cached_data else set()
        
        resp.raise_for_status()
        
        # ページネーション処理
        data, new_etag = paginate_github_api(url, token, etag=cached_etag)
        members = {m.get("login") for m in data if m.get("login")}
        logging.info("チームメンバー %d 名を取得しました", len(members))
        
        # キャッシュ保存（setはJSON化できないのlistに変換）
        save_cache(cache_key, list(members), new_etag)
        
        return members
        
    except requests.exceptions.RequestException as e:
        logging.error("チームメンバー取得エラー: %s", e)
        return set()


# =========================
# CSV 読み込み
# =========================
def read_user_csv(path: str) -> List[Dict[str, str]]:
    """ユーザー一覧 CSV を読み込む"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# =========================
# メイン処理
# =========================
def main() -> None:
    if len(sys.argv) < 4:
        print("使い方: python copilot_summary_append.py user.csv .env output.xlsx", file=sys.stderr)
        sys.exit(1)

    user_csv_path = sys.argv[1]
    config_path = sys.argv[2]
    summary_path = sys.argv[3]

    # 設定ロード
    logging.info("設定ファイルを読み込み中: %s", config_path)
    token, org, team_slug = load_config_from_env_file(config_path)

    # チームメンバー取得（TARGET_TEAM_SLUG が設定されている場合のみ）
    team_members: Optional[Set[str]] = None
    if team_slug:
        team_members = fetch_team_members(org, token, team_slug)
        print(f"TARGET_TEAM_SLUG={team_slug} が指定されました。チームメンバー {len(team_members)} 名を対象とします。")

    # ユーザー CSV 読み込み
    logging.info("ユーザー CSV を読み込み中: %s", user_csv_path)
    users = read_user_csv(user_csv_path)
    logging.info("ユーザー数: %d 名", len(users))

    # Copilot シート情報取得
    seats = fetch_all_seats(org, token)
    seat_by_login: Dict[str, Dict[str, Any]] = {
        seat.get("assignee", {}).get("login"): seat
        for seat in seats
        if seat.get("assignee", {}).get("login")
    }

    now_utc = datetime.now(timezone.utc)

    total_users = 0
    inactive_count = 0

    logging.info("利用状況を判定中...")
    # 対象ユーザーごとに「%d日以上未利用かどうか」を判定
    for u in users:
        login = u.get("GitHubユーザー名", "")

        # チーム指定がある場合は、team_members に含まれないユーザーはスキップ
        if team_members is not None and login not in team_members:
            continue

        total_users += 1

        seat = seat_by_login.get(login)
        inactive = False

        if seat:
            raw_last = seat.get("last_activity_at") or ""
            last_dt = parse_iso_datetime(raw_last)

            if last_dt is None:
                inactive = True
            else:
                delta = now_utc - last_dt.astimezone(timezone.utc)
                if delta.days >= INACTIVE_DAYS:
                    inactive = True
        else:
            # シート自体が無い場合も「未利用」とみなす
            inactive = True

        if inactive:
            inactive_count += 1

    active_count = max(total_users - inactive_count, 0)

    if total_users > 0:
        inactive_pct = f"{round(inactive_count / total_users * 100)}%"
        active_pct = f"{round(active_count / total_users * 100)}%"
    else:
        inactive_pct = ""
        active_pct = ""

    # JST の現在時刻
    jst = timezone(timedelta(hours=JST_OFFSET))
    now_jst = datetime.now(jst).strftime("%Y/%m/%d %H:%M")

    logging.info("統計: 総ユーザー数=%d, %d日以上未利用=%d, %d日以内利用=%d", 
                 total_users, INACTIVE_DAYS, inactive_count, INACTIVE_DAYS, active_count)

    # Excel に追記
    if os.path.exists(summary_path):
        logging.info("既存のサマリーファイルを読み込み中: %s", summary_path)
        wb = load_workbook(summary_path)
        ws = wb.active
    else:
        logging.info("新規サマリーファイルを作成中: %s", summary_path)
        wb = Workbook()
        ws = wb.active
        ws.append([
            "データ取得日時",
            f"{INACTIVE_DAYS}日以上未利用ユーザー数",
            f"{INACTIVE_DAYS}日以上未利用ユーザー数(%)",
            f"{INACTIVE_DAYS}日以内に利用したユーザー数",
            f"{INACTIVE_DAYS}日以内に利用したユーザー数(%)",
        ])

    ws.append([
        now_jst,
        inactive_count,
        inactive_pct,
        active_count,
        active_pct,
    ])

    try:
        logging.info("サマリーファイルを保存中: %s", summary_path)
        wb.save(summary_path)
        logging.info("保存完了")
    except PermissionError as e:
        logging.error("サマリーファイル '%s' を保存できませんでした。", summary_path)
        logging.error("Excel などでファイルが開かれていないか確認し、閉じてからもう一度実行してください。")
        logging.error("詳細: %s", e)
        sys.exit(1)
    except (OSError, IOError) as e:
        logging.error("予期しないエラーが発生しました: %s", e)
        sys.exit(1)

    print(f"サマリーを追記しました: {summary_path}")
    print(f"対象ユーザー数: {total_users}, {INACTIVE_DAYS}日以上未利用: {inactive_count}, {INACTIVE_DAYS}日以内利用: {active_count}")


if __name__ == "__main__":
    main()
